import json
import logging
import os
import re
import sqlite3
import unicodedata
from io import BytesIO
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, HTTPException, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, ConfigDict, EmailStr, Field
import requests

try:
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import DateRange, Dimension, Metric, RunReportRequest
    from google.oauth2 import service_account

    GA4_CLIENT_AVAILABLE = True
except ImportError:
    GA4_CLIENT_AVAILABLE = False


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger("ai-portfolio-inbox")
CHAT_WIDGET_SOURCE = "portfolio-chat-widget"
WHATSAPP_NUMBER = "+34 691068400"
WHATSAPP_LINK = "https://wa.me/34691068400"
WHATSAPP_TEXT = {
    "es": "Hola Carlos, te contacto desde tu portfolio.",
    "en": "Hi Carlos, I am contacting you from your portfolio.",
}


LANGUAGE_STOPWORDS = {
    "es": {"hola", "gracias", "quiero", "necesito", "proyecto", "precio", "colaboracion", "colaboración", "para", "una", "con", "como", "puedes", "sugerencia", "error", "sitio", "mensaje"},
    "en": {"hello", "thanks", "project", "pricing", "budget", "question", "suggestion", "issue", "website", "message", "need", "would", "like", "with", "about", "feature", "collaboration"},
}

PRIORITY_RANK = {"low": 1, "medium": 2, "high": 3}

THEME_RULES = [
    ("ai-automation", "AI automation and assistant builds", {"ai", "automation", "assistant", "agents", "chatbot", "llm", "openai", "copilot", "automatizacion"}),
    ("project-collaboration", "Project collaboration opportunities", {"project", "projects", "hire", "freelance", "contract", "collaboration", "partnership", "consulting", "proyecto", "colaboracion", "colaboración"}),
    ("pricing-budget", "Pricing and budget discussions", {"pricing", "budget", "quote", "cost", "proposal", "price", "precio", "presupuesto", "cotizacion", "cotización"}),
    ("bug-performance", "Bugs and performance issues", {"bug", "broken", "error", "issue", "problem", "slow", "crash", "fix", "fallo", "problema", "lento"}),
    ("analytics-dashboard", "Analytics and dashboard requests", {"analytics", "dashboard", "ga4", "google", "tracking", "reporting", "insights", "metricas", "métricas"}),
    ("content-portfolio", "Portfolio content and case studies", {"portfolio", "case", "study", "content", "copy", "about", "resume", "cv", "portafolio", "contenido"}),
    ("ux-feedback", "Portfolio UX and design feedback", {"feedback", "suggestion", "improve", "design", "ui", "ux", "layout", "experience", "sugerencia", "mejora"}),
    ("api-integration", "API and integration topics", {"api", "integration", "webhook", "backend", "database", "sqlite", "fastapi", "smtp", "integracion", "integración"}),
]

CHAT_PROFILE_FACTS = {
    "en": [
        "Carlos works at the intersection of operations, data, digital workflows, and practical AI.",
        "Carlos currently works at The Retail Performance Company (RPC) since November 2025.",
        "His profile is corporate, business-oriented, and grounded in operations, reporting, process coordination, and digital support.",
        "At RPC, his work is focused on operations, reporting, process support, digital workflows, and practical AI in a corporate environment linked to BMW-related processes.",
        "He has experience supporting Purchasing and Aftersales processes in a BMW-related environment, including SAP support, reporting follow-up, incident handling, and coordination.",
        "His background also includes back-office operations, documentation validation, public procurement support, banking operations, and technical support operations.",
        "His projects show practical uses of digital tools, dashboards, structured information, and AI for productivity.",
    ],
    "es": [
        "Carlos trabaja en la interseccion entre operaciones, datos, workflows digitales e IA practica.",
        "Carlos trabaja actualmente en The Retail Performance Company (RPC) desde noviembre de 2025.",
        "Su perfil es corporativo, orientado a negocio y centrado en operaciones, reporting, coordinacion de procesos y soporte digital.",
        "En RPC, su trabajo se centra en operaciones, reporting, soporte a procesos, workflows digitales e IA practica en un entorno corporativo relacionado con procesos vinculados a BMW.",
        "Tiene experiencia dando soporte a procesos de Purchasing y Aftersales en un entorno vinculado a BMW, incluyendo soporte SAP, seguimiento de reporting, gestion de incidencias y coordinacion.",
        "Su trayectoria tambien incluye back office, validacion documental, soporte a contratacion publica, operaciones bancarias y soporte tecnico-operativo.",
        "Sus proyectos muestran usos practicos de herramientas digitales, dashboards, informacion estructurada e IA para productividad.",
    ],
}

PORTFOLIO_ROLE_FACTS = {
    "en": {
        "current_role": "Carlos currently works at The Retail Performance Company (RPC) since November 2025 as an Operations & Business Support Consultant. His current work is focused on Purchasing and Aftersales support, SAP-related follow-up, reporting, incident handling, and operational coordination in a BMW-related business environment.",
        "education": [
            "Bachelor's Degree in Business Administration at Universidad Rey Juan Carlos, with an academic foundation in business management, organizational processes, finance, and operational decision-making.",
            "Higher Technician in Network Systems Administration (ASIR) at Universidad Europea, focused on systems administration, infrastructure fundamentals, and IT support environments.",
            "University continuing studies in Cybersecurity, Artificial Intelligence and Big Data at Universidad Europea, reinforcing analytical and digital capabilities applied to operations contexts.",
        ],
        "previous_experience": [
            "Endesa (2025) - Back Office Operations, focused on documentation, validation, and coordination in solar financing workflows using Salesforce.",
            "Ayuntamiento de Madrid (2022 - 2025) - Administrative Procurement Support, with work in documentation control, public procurement coordination, and process follow-up.",
            "Openbank (2021 - 2022) - Banking Customer Support, covering customer operations, fraud alerts, KYC checks, and regulated banking workflows.",
            "Movistar Prosegur Alarmas (2019 - 2020) - Technical Support Operations, combining incident handling, service coordination, and technician scheduling.",
        ],
    },
    "es": {
        "current_role": "Carlos trabaja actualmente en The Retail Performance Company (RPC) desde noviembre de 2025 como Consultor de Operaciones y Soporte de Negocio. Su trabajo actual se centra en soporte a Purchasing y Aftersales, seguimiento relacionado con SAP, reporting, gestion de incidencias y coordinacion operativa en un entorno de negocio vinculado a BMW.",
        "education": [
            "Grado en Administracion y Direccion de Empresas en la Universidad Rey Juan Carlos, con base academica en gestion empresarial, procesos organizativos, finanzas y toma de decisiones operativas.",
            "Tecnico Superior en Administracion de Sistemas en Red (ASIR) en la Universidad Europea, con formacion en administracion de sistemas, fundamentos de infraestructura y entornos de soporte TI.",
            "Estudios universitarios complementarios en Ciberseguridad, Inteligencia Artificial y Big Data en la Universidad Europea, orientados a reforzar capacidades analiticas y digitales aplicadas a entornos operativos.",
        ],
        "previous_experience": [
            "Endesa (2025) - Operaciones Back Office, con foco en documentacion, validacion y coordinacion en workflows de financiacion solar usando Salesforce.",
            "Ayuntamiento de Madrid (2022 - 2025) - Soporte Administrativo en Contratacion, trabajando en control documental, coordinacion de contratacion publica y seguimiento de procesos.",
            "Openbank (2021 - 2022) - Soporte al Cliente Bancario, cubriendo operativa de cliente, alertas de fraude, controles KYC y workflows bancarios regulados.",
            "Movistar Prosegur Alarmas (2019 - 2020) - Operaciones de Soporte Tecnico, combinando gestion de incidencias, coordinacion de servicio y planificacion de visitas de tecnicos.",
        ],
    },
}

EXPERIENCE_DETAIL_FACTS = {
    "rpc": {
        "terms": {"rpc", "retail performance company"},
        "en": "At The Retail Performance Company (RPC), Carlos works as an Operations & Business Support Consultant. The role combines Purchasing and Aftersales support, SAP-related follow-up, reporting, incident handling, coordination, and data visibility in a BMW-related business environment.",
        "es": "En The Retail Performance Company (RPC), Carlos trabaja como Consultor de Operaciones y Soporte de Negocio. El rol combina soporte a Purchasing y Aftersales, seguimiento relacionado con SAP, reporting, gestion de incidencias, coordinacion y visibilidad de datos en un entorno de negocio vinculado a BMW.",
        "thread_title": "RPC experience",
    },
    "endesa": {
        "terms": {"endesa"},
        "en": "At Endesa, Carlos worked in back-office operations focused on documentation, validation, and coordination in solar financing workflows. He also used Salesforce to track incidents, updates, and process status.",
        "es": "En Endesa, Carlos trabajo en operaciones de back office centradas en documentacion, validacion y coordinacion dentro de workflows de financiacion solar. Tambien utilizaba Salesforce para registrar incidencias, actualizaciones y el estado del proceso.",
        "thread_title": "Endesa experience",
    },
    "ayuntamiento-madrid": {
        "terms": {"ayuntamiento de madrid", "madrid city council", "ayuntamiento", "contratacion publica"},
        "en": "At Ayuntamiento de Madrid, Carlos provided administrative support in public procurement. His work was centered on documentation control, tender and contract records, platform handling, coordination, and process follow-up across the different administrative steps.",
        "es": "En el Ayuntamiento de Madrid, Carlos dio soporte administrativo en contratacion publica. Su trabajo se centraba en control documental, gestion de expedientes y licitaciones, uso de plataformas de contratacion, coordinacion y seguimiento de las distintas fases administrativas.",
        "thread_title": "Ayuntamiento de Madrid experience",
    },
    "openbank": {
        "terms": {"openbank"},
        "en": "At Openbank, Carlos worked in operational banking support. The role included customer queries, fraud-alert tasks, KYC checks, and support for regulated banking workflows with attention to accuracy and response time.",
        "es": "En Openbank, Carlos trabajo en soporte bancario operativo. El rol incluia atencion a consultas de clientes, tareas relacionadas con alertas de fraude, controles KYC y soporte a workflows bancarios regulados con foco en precision y tiempos de respuesta.",
        "thread_title": "Openbank experience",
    },
    "movistar-prosegur": {
        "terms": {"movistar prosegur alarmas", "prosegur alarmas", "movistar"},
        "en": "At Movistar Prosegur Alarmas, Carlos worked in technical support operations. The role combined incident handling, service coordination, and scheduling or follow-up for technician visits.",
        "es": "En Movistar Prosegur Alarmas, Carlos trabajo en operaciones de soporte tecnico. El rol combinaba gestion de incidencias, coordinacion del servicio y planificacion o seguimiento de visitas de tecnicos.",
        "thread_title": "Movistar Prosegur experience",
    },
}

TOOLING_FACTS = {
    "en": {
        "current_tools": "In his current role at The Retail Performance Company (RPC), Carlos mainly works with SAP, Excel, and Qlik Sense.",
        "previous_tools": "In previous experience, the clearest tool explicitly reflected in the portfolio is Salesforce, especially in Endesa.",
        "sap": "Yes. Carlos works with SAP in his current role at The Retail Performance Company (RPC).",
        "excel": "Yes. Carlos uses Excel in his current role at The Retail Performance Company (RPC).",
        "qlik-sense": "Yes. Carlos uses Qlik Sense in his current role at The Retail Performance Company (RPC).",
        "salesforce": "Yes. Carlos used Salesforce especially in his experience at Endesa.",
        "endesa_tools": "At Endesa, the most relevant tool reflected in the portfolio is Salesforce.",
    },
    "es": {
        "current_tools": "En su trabajo actual en The Retail Performance Company (RPC), Carlos utiliza principalmente SAP, Excel y Qlik Sense.",
        "previous_tools": "En experiencia previa, la herramienta mas clara que aparece reflejada en el portfolio es Salesforce, especialmente en Endesa.",
        "sap": "Si. Carlos trabaja con SAP en su rol actual en The Retail Performance Company (RPC).",
        "excel": "Si. Carlos utiliza Excel en su rol actual en The Retail Performance Company (RPC).",
        "qlik-sense": "Si. Carlos utiliza Qlik Sense en su rol actual en The Retail Performance Company (RPC).",
        "salesforce": "Si. Carlos ha usado Salesforce especialmente en su experiencia en Endesa.",
        "endesa_tools": "En Endesa, la herramienta mas relevante que aparece reflejada en el portfolio es Salesforce.",
    },
}

PROJECT_SPOTLIGHTS = {
    "en": [
        {
            "title": "Dashboards, KPI Tracking and Data Analysis",
            "summary": "Practical dashboard and reporting work focused on KPI visibility, operational follow-up, and clearer business monitoring.",
            "value": "It matters because it shows how Carlos turns data into usable visibility for teams and day-to-day decisions.",
        },
        {
            "title": "AI Tools for Content and Productivity",
            "summary": "Applied use of AI tools to organize information, support writing, and simplify recurring tasks.",
            "value": "It matters because it reflects a realistic, productivity-focused use of practical AI in business routines.",
        },
        {
            "title": "Personal Portfolio Website",
            "summary": "A structured digital project that brings together experience, projects, and professional communication in one place.",
            "value": "It matters because it shows how Carlos organizes information and presents workflows, experience, and digital initiatives clearly.",
        },
    ],
    "es": [
        {
            "title": "Dashboards, KPI y Analisis de Datos",
            "summary": "Trabajo practico de dashboards y reporting orientado a visibilidad KPI, seguimiento operativo y lectura mas clara del negocio.",
            "value": "Importa porque muestra como Carlos convierte datos en visibilidad util para equipos y decisiones del dia a dia.",
        },
        {
            "title": "Herramientas de IA para Contenido y Productividad",
            "summary": "Uso aplicado de herramientas de IA para organizar informacion, apoyar redaccion y simplificar tareas recurrentes.",
            "value": "Importa porque refleja un uso realista y orientado a productividad de la IA practica en rutinas de negocio.",
        },
        {
            "title": "Portfolio Personal",
            "summary": "Proyecto digital estructurado que reune experiencia, proyectos y comunicacion profesional en un solo espacio.",
            "value": "Importa porque muestra como Carlos organiza informacion y presenta con claridad workflows, experiencia e iniciativas digitales.",
        },
    ],
}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def iso_days_ago(days: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=days)).date().isoformat()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "general"


def parse_json_list(raw: Any, fallback: list[str] | None = None) -> list[str]:
    if isinstance(raw, list):
        return [str(item) for item in raw]
    if not raw:
        return fallback or []
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return fallback or []
    if isinstance(parsed, list):
        return [str(item) for item in parsed]
    return fallback or []


def normalize_list(values: list[str], limit: int = 6) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = re.sub(r"\s+", " ", str(value)).strip()
        if text and text.lower() not in seen:
            seen.add(text.lower())
            cleaned.append(text)
    return cleaned[:limit]


def tokenize(text: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-zA-Záéíóúñü0-9]{3,}", text.lower())
        if token not in {"this", "that", "with", "from", "have", "your", "about", "would", "para", "como", "esto", "esta", "este"}
    }


def detect_language(text: str) -> str:
    tokens = tokenize(text)
    es_score = len(tokens & LANGUAGE_STOPWORDS["es"])
    en_score = len(tokens & LANGUAGE_STOPWORDS["en"])
    if re.search(r"[ñáéíóú]", text.lower()):
        es_score += 2
    return "es" if es_score > en_score else "en"


def normalize_match_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_text).strip().lower()


def contains_standalone_term(text: str, term: str) -> bool:
    return re.search(rf"(?<![a-z0-9]){re.escape(term)}(?![a-z0-9])", text) is not None


def build_whatsapp_url(language: str) -> str:
    message = WHATSAPP_TEXT.get(language, WHATSAPP_TEXT["en"])
    return f"{WHATSAPP_LINK}?text={quote(message)}"


def language_scores(text: str) -> tuple[int, int]:
    normalized = normalize_match_text(text)
    if not normalized:
        return 0, 0

    tokens = set(re.findall(r"[a-z0-9]{2,}", normalized))
    es_terms = {
        "hola", "gracias", "quiero", "necesito", "puedes", "podrias", "perfil", "resumen", "proyectos",
        "experiencia", "contacto", "contactar", "encaje", "operaciones", "proceso", "procesos",
        "datos", "mostrar", "muestrame", "como", "por", "que", "hablame", "cuentame", "espanol",
        "ahora", "sigue", "seguimos", "mas", "relevantes", "rol", "roles",
    }
    en_terms = {
        "hello", "thanks", "want", "need", "can", "could", "profile", "summary", "projects",
        "experience", "contact", "fit", "operations", "process", "processes", "data", "show",
        "about", "why", "how", "english", "continue", "relevant", "role", "roles", "please",
    }
    es_score = len(tokens & es_terms)
    en_score = len(tokens & en_terms)

    if any(marker in normalized for marker in {" que ", " como ", " para ", " con ", " el ", " la ", " los ", " las "}):
        es_score += 1
    if any(marker in normalized for marker in {" the ", " and ", " with ", " for ", " his ", " her ", " role ", " roles "}):
        en_score += 1
    if re.search(r"[¿¡]", text):
        es_score += 2
    if re.search(r"[áéíóúñÁÉÍÓÚÑ]", text):
        es_score += 2

    return es_score, en_score


def explicit_chat_language(normalized_text: str) -> str | None:
    english_switches = {
        "answer in english", "respond in english", "reply in english", "continue in english",
        "can you continue in english", "speak in english", "english please", "en ingles",
        "puedes seguir en ingles", "ahora en ingles", "respondeme en ingles",
    }
    spanish_switches = {
        "answer in spanish", "respond in spanish", "reply in spanish", "continue in spanish",
        "can you continue in spanish", "speak in spanish", "spanish please", "en espanol",
        "respondeme en espanol", "ahora en espanol", "puedes seguir en espanol",
    }
    if any(term in normalized_text for term in english_switches):
        return "en"
    if any(term in normalized_text for term in spanish_switches):
        return "es"
    return None


def detect_language_from_history(messages: list[dict[str, str]] | None) -> str | None:
    if not messages:
        return None

    recent_user_messages = [
        item.get("content", "").strip()
        for item in reversed(messages[-6:])
        if item.get("role") == "user" and item.get("content")
    ]
    if not recent_user_messages:
        return None

    es_total = 0
    en_total = 0
    for content in recent_user_messages[:3]:
        normalized = normalize_match_text(content)
        explicit = explicit_chat_language(normalized)
        if explicit:
            return explicit
        es_score, en_score = language_scores(content)
        es_total += es_score
        en_total += en_score

    if es_total == en_total:
        return None
    return "es" if es_total > en_total else "en"


def resolve_chat_language_details(submission: "InboxSubmission") -> tuple[str, str]:
    normalized_message = normalize_match_text(submission.message)
    explicit = explicit_chat_language(normalized_message)
    if explicit:
        return explicit, f"explicit_{'spanish' if explicit == 'es' else 'english'}"

    current_es, current_en = language_scores(submission.message)
    if current_es != current_en and max(current_es, current_en) >= 1:
        return ("es", "user_message") if current_es > current_en else ("en", "user_message")

    history_language = detect_language_from_history(submission.messages)
    if history_language:
        return history_language, "conversation_history"

    if submission.locale in {"es", "en"}:
        return submission.locale, "frontend_locale"

    return detect_language(submission.message), "legacy_detector"


def resolve_chat_language(submission: "InboxSubmission") -> str:
    language, _ = resolve_chat_language_details(submission)
    return language


@dataclass
class Settings:
    database_path: Path
    openai_api_key: str
    openai_model: str
    resend_api_key: str
    resend_from_email: str
    smtp_host: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    smtp_use_tls: bool
    email_from: str
    owner_email: str
    app_base_url: str
    ai_match_threshold: float
    ga4_property_id: str
    google_application_credentials: str
    google_application_credentials_json: str


def parse_bool_env(value: str, default: bool = False) -> bool:
    normalized = (value or "").strip().lower()
    if not normalized:
        return default
    return normalized in {"1", "true", "yes", "on"}


def parse_int_env(name: str, default: int) -> int:
    raw = os.getenv(name, str(default)).strip()
    try:
        return int(raw)
    except ValueError:
        logger.warning("Invalid integer for %s=%r. Falling back to %s.", name, raw, default)
        return default


def resolve_database_path() -> Path:
    configured = os.getenv("DATABASE_PATH", "").strip()
    if configured:
        candidate = Path(configured)
        return candidate if candidate.is_absolute() else BASE_DIR / candidate

    render_disk_root = os.getenv("RENDER_DISK_ROOT", "").strip()
    if render_disk_root:
        return Path(render_disk_root) / "ai_portfolio_inbox.db"

    return BASE_DIR / "ai_portfolio_inbox.db"


def load_settings() -> Settings:
    owner_email = os.getenv("OWNER_EMAIL", "c.sanmiguelortega@gmail.com").strip()
    smtp_username = os.getenv("SMTP_USERNAME", "").strip()
    email_from = os.getenv("EMAIL_FROM", "").strip() or smtp_username
    resend_from_email = os.getenv("RESEND_FROM_EMAIL", "").strip() or email_from
    return Settings(
        database_path=resolve_database_path(),
        openai_api_key=os.getenv("OPENAI_API_KEY", "").strip(),
        openai_model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini").strip(),
        resend_api_key=os.getenv("RESEND_API_KEY", "").strip(),
        resend_from_email=resend_from_email,
        smtp_host=os.getenv("SMTP_HOST", "").strip(),
        smtp_port=parse_int_env("SMTP_PORT", 587),
        smtp_username=smtp_username,
        smtp_password=os.getenv("SMTP_PASSWORD", "").strip(),
        smtp_use_tls=parse_bool_env(os.getenv("SMTP_USE_TLS", "true"), default=True),
        email_from=email_from,
        owner_email=os.getenv("OWNER_EMAIL", owner_email).strip() or owner_email,
        app_base_url=os.getenv("APP_BASE_URL", "http://127.0.0.1:8000").strip(),
        ai_match_threshold=float(os.getenv("AI_MATCH_THRESHOLD", "0.33")),
        ga4_property_id=os.getenv("GA4_PROPERTY_ID", "").strip(),
        google_application_credentials=os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip(),
        google_application_credentials_json=(
            os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON", "").strip()
            or os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
        ),
    )


settings = load_settings()

# Gmail SMTP recommended configuration:
# SMTP_HOST=smtp.gmail.com
# SMTP_PORT=587
# SMTP_USE_TLS=true
# SMTP_USERNAME=mi_correo_gmail
# SMTP_PASSWORD=app_password_de_google
# EMAIL_FROM=mi_correo_gmail
# RESEND_API_KEY=re_xxx
# RESEND_FROM_EMAIL=Portfolio Inbox <onboarding@resend.dev>
# OWNER_EMAIL=mi_correo_personal_destino

DEFAULT_CORS_ORIGINS = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "https://portfolio-khaki-zeta-3frz86na3s.vercel.app",
]
DEFAULT_CORS_ORIGIN_REGEX = r"^https:\/\/[a-z0-9-]+\.vercel\.app$"
CORS_ALLOW_HEADERS = [
    "Accept",
    "Authorization",
    "Content-Type",
    "Origin",
    "X-Requested-With",
    "x-chat-locale",
    "x-chat-source",
]


def normalize_origin(value: str) -> str:
    return value.strip().rstrip("/")


class InboxSubmission(BaseModel):
    name: str | None = Field(default=None, max_length=80)
    email: str | None = Field(default=None, max_length=320)
    company: str | None = Field(default=None, max_length=120)
    message: str | None = Field(default=None, max_length=3000)
    source: str | None = Field(default=None, max_length=80)
    messages: list[dict[str, Any]] | None = None
    history: list[dict[str, Any]] | None = None
    locale: str | None = Field(default=None, max_length=16)

    model_config = ConfigDict(str_strip_whitespace=True, extra="allow")


class MessageAnalysis(BaseModel):
    language: str = Field(pattern="^(es|en)$")
    category: str
    priority: str
    summary: str
    key_points: list[str]
    theme_label: str
    theme_slug: str
    thread_title: str
    reply_text: str
    lead_score: int = Field(ge=1, le=5)
    sentiment: str


EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def clean_text(value: Any, limit: int, default: str = "") -> str:
    if value is None:
        return default
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return default
    return text[:limit]


def clean_optional_text(value: Any, limit: int) -> str | None:
    text = clean_text(value, limit)
    return text or None


def clean_email(value: Any) -> str | None:
    candidate = clean_text(value, 320)
    if not candidate:
        return None
    if EMAIL_REGEX.match(candidate):
        return candidate
    logger.warning("Ignoring invalid email value in inbox payload")
    return None


def normalize_locale_value(value: Any) -> str | None:
    candidate = clean_text(value, 16).lower()
    if candidate.startswith("es"):
        return "es"
    if candidate.startswith("en"):
        return "en"
    return None


def normalize_history_messages(value: Any) -> list[dict[str, str]]:
    if not isinstance(value, list):
        return []

    normalized_items: list[dict[str, str]] = []
    for item in value[:12]:
        if not isinstance(item, dict):
            continue
        role = clean_text(item.get("role"), 20).lower()
        content = clean_text(item.get("content") or item.get("message") or item.get("text"), 1200)
        if role not in {"user", "assistant", "system"} or not content:
            continue
        normalized_items.append({"role": role, "content": content})
    return normalized_items


def safe_payload_dict(raw_payload: Any) -> dict[str, Any]:
    return raw_payload if isinstance(raw_payload, dict) else {}


def build_default_analysis(message: str, language: str, source: str) -> MessageAnalysis:
    preview = message[:120] if message else ("Mensaje recibido" if language == "es" else "Message received")
    source_label = source or "portfolio-chat-widget"
    summary = (
        "Consulta recibida correctamente. El sistema ha usado una respuesta de respaldo para mantener la conversacion activa."
        if language == "es"
        else "Message received successfully. The system used a fallback reply to keep the conversation moving."
    )
    return MessageAnalysis(
        language=language,
        category="general feedback",
        priority="medium" if source_label == CHAT_WIDGET_SOURCE else "low",
        summary=summary,
        key_points=[preview] if preview else [],
        theme_label="General inbound inquiries",
        theme_slug="general-inquiries",
        thread_title=preview or ("Consulta general" if language == "es" else "General inquiry"),
        reply_text="",
        lead_score=2,
        sentiment="neutral",
    )


def build_locale_fallback_reply(language: str) -> str:
    if language == "es":
        return "Lo siento, ha ocurrido un problema. Puedes contactar directamente con Carlos por WhatsApp."
    return "Sorry, something went wrong. You can contact Carlos directly via WhatsApp."


def normalize_inbox_submission(raw_payload: Any, request: Request) -> tuple[InboxSubmission, str]:
    payload = safe_payload_dict(raw_payload)
    history = normalize_history_messages(payload.get("messages"))
    if not history:
        history = normalize_history_messages(payload.get("history"))

    message = clean_text(payload.get("message"), 3000)
    if not message and history:
        last_user_message = next(
            (item["content"] for item in reversed(history) if item.get("role") == "user" and item.get("content")),
            "",
        )
        message = clean_text(last_user_message, 3000)

    body_locale = normalize_locale_value(payload.get("locale"))
    header_locale = normalize_locale_value(request.headers.get("x-chat-locale"))
    locale = body_locale or header_locale

    normalized_payload = InboxSubmission.model_validate(
        {
            "name": clean_optional_text(payload.get("name"), 80) or "Website Visitor",
            "email": clean_email(payload.get("email")),
            "company": clean_optional_text(payload.get("company"), 120),
            "message": message or "",
            "source": clean_optional_text(payload.get("source"), 80) or CHAT_WIDGET_SOURCE,
            "messages": history,
            "locale": locale,
        }
    )
    return normalized_payload, header_locale or ""


def build_message_payload(
    submission: InboxSubmission,
    analysis: MessageAnalysis,
    engine: str,
    reply_text: str,
    email_status: str,
    saved: bool,
    message_id: int | None = None,
    thread_id: int | None = None,
    thread_title: str | None = None,
    chat_meta: dict[str, Any] | None = None,
    created_at: str | None = None,
) -> dict[str, Any]:
    meta = chat_meta or {}
    whatsapp_handoff = bool(meta.get("whatsapp_handoff", False))
    whatsapp_url = build_whatsapp_url(analysis.language) if whatsapp_handoff else ""
    return {
        "id": message_id,
        "thread_id": thread_id,
        "thread_title": thread_title or analysis.thread_title,
        "theme_slug": analysis.theme_slug,
        "theme_label": analysis.theme_label,
        "user_name": submission.name,
        "user_email": submission.email,
        "company": submission.company,
        "source": submission.source or CHAT_WIDGET_SOURCE,
        "language": analysis.language,
        "category": analysis.category,
        "priority": analysis.priority,
        "lead_score": analysis.lead_score,
        "sentiment": analysis.sentiment,
        "summary": analysis.summary,
        "key_points": analysis.key_points,
        "raw_message": submission.message or "",
        "reply_text": reply_text,
        "thread_summary": "",
        "email_status": allowed_email_status(email_status),
        "analysis_engine": engine,
        "feedback_signal": meta.get("feedback_signal", "none"),
        "feedback_reason": meta.get("feedback_reason") or "",
        "chat_intent": meta.get("intent") or "",
        "whatsapp_handoff": whatsapp_handoff,
        "whatsapp_url": whatsapp_url,
        "created_at": created_at or utc_now(),
        "saved": saved,
    }


def finalize_message_notification(
    message_id: int,
    submission: "InboxSubmission",
    analysis: "MessageAnalysis",
    thread: dict[str, Any],
    thread_id: int | None,
    is_chat_request: bool,
) -> None:
    related_messages: list[dict[str, Any]] = []
    email_status = "skipped"

    try:
        with closing(get_connection()) as connection:
            if thread_id is not None:
                try:
                    related_messages = get_thread_messages(connection, thread_id, limit=4)
                except Exception:
                    logger.exception("Loading related thread messages failed inside notification task")
                    related_messages = []

            logger.info(
                "Inbox email trigger starting | message_id=%s | thread_id=%s | source=%s | chat_request=%s",
                message_id,
                thread_id,
                submission.source,
                is_chat_request,
            )
            email_status = allowed_email_status(
                send_email_notification(message_id, submission, analysis, thread or {}, related_messages)
            )
            logger.info("Inbox email trigger finished | message_id=%s | email_status=%s", message_id, email_status)

            connection.execute("UPDATE messages SET email_status = ? WHERE id = ?", (email_status, message_id))
            connection.commit()
            logger.info("Inbox email status updated | message_id=%s | email_status=%s", message_id, email_status)
    except Exception:
        logger.exception("Notification task failed | message_id=%s | source=%s", message_id, submission.source)


async def handle_inbox_submission(request: Request, background_tasks: BackgroundTasks | None = None) -> JSONResponse:
    raw_payload: dict[str, Any] = {}
    submission: InboxSubmission | None = None
    final_language = "en"
    language_source = "default"
    is_chat_request = False
    engine = "fallback"
    email_status = "skipped"
    chat_meta: dict[str, Any] = {
        "feedback_signal": "none",
        "feedback_reason": None,
        "intent": "",
        "whatsapp_handoff": False,
    }
    thread: dict[str, Any] | None = None
    related_messages: list[dict[str, Any]] = []
    saved = False
    message_id: int | None = None
    thread_id: int | None = None
    thread_title: str | None = None
    created_at = utc_now()

    try:
        raw_payload = await request.json()
    except Exception:
        logger.warning("Inbox request body could not be parsed as JSON; continuing with empty payload")
        raw_payload = {}

    try:
        submission, header_locale = normalize_inbox_submission(raw_payload, request)
        is_chat_request = submission.source == CHAT_WIDGET_SOURCE or bool(submission.messages)
        if is_chat_request:
            final_language, language_source = resolve_chat_language_details(submission)
        else:
            final_language = submission.locale or header_locale or detect_language(submission.message or "")
            language_source = "body_header_or_message"
        submission = submission.model_copy(update={"locale": final_language})
    except Exception:
        logger.exception("Inbox input normalization failed; using emergency defaults")
        emergency_locale = normalize_locale_value(request.headers.get("x-chat-locale")) or "en"
        submission = InboxSubmission(
            name="Website Visitor",
            email=None,
            company=None,
            message="",
            source=CHAT_WIDGET_SOURCE,
            messages=[],
            locale=emergency_locale,
        )
        final_language = emergency_locale
        language_source = "emergency_default"
        is_chat_request = True

    assert submission is not None
    logger.info(
        "Inbox submission received | source=%s | locale=%s | language=%s | language_source=%s | has_history=%s | email_present=%s | message_length=%s",
        submission.source,
        submission.locale or "auto",
        final_language,
        language_source,
        bool(submission.messages),
        bool(submission.email),
        len(submission.message or ""),
        extra={
            "email": submission.email,
            "company": submission.company,
            "source": submission.source,
            "message_preview": (submission.message or "")[:160],
        },
    )

    reply_text = build_locale_fallback_reply(final_language)
    analysis = build_default_analysis(submission.message or "", final_language, submission.source or CHAT_WIDGET_SOURCE)

    if not submission.message:
        reply_text = (
            "Puedo ayudarte con perfil, experiencia, proyectos o contacto. Si quieres, escribe tu pregunta en una frase y sigo desde ahi."
            if final_language == "es"
            else "I can help with profile, experience, projects, or contact. If you want, send your question in one sentence and I will take it from there."
        )
        analysis = analysis.model_copy(
            update={
                "summary": "Empty or partial payload received; a safe reply was returned.",
                "reply_text": reply_text,
            }
        )
        engine = "input-fallback"
    else:
        try:
            if is_chat_request:
                logger.info(
                    "Chat request trace | endpoint=/api/inbox | source=%s | chat_locale=%s | chat_language=%s | chat_language_source=%s | history_items=%s",
                    submission.source,
                    submission.locale or "none",
                    final_language,
                    language_source,
                    len(submission.messages or []),
                )
                analysis, _ = heuristic_analysis(submission)
                reply_text, engine, chat_meta = generate_chat_reply(submission, analysis, final_language)
                analysis = analysis.model_copy(update={"language": final_language, "reply_text": reply_text})
                logger.info(
                    "Chat response trace | chat_response_path=%s | chat_language=%s | feedback_signal=%s | feedback_reason=%s | chat_intent=%s | whatsapp_handoff=%s",
                    engine,
                    final_language,
                    chat_meta["feedback_signal"],
                    chat_meta["feedback_reason"] or "none",
                    chat_meta["intent"],
                    chat_meta["whatsapp_handoff"],
                )
            else:
                analysis, engine = openai_analysis(submission)
                final_language = analysis.language
                reply_text = analysis.reply_text or build_locale_fallback_reply(final_language)
                analysis = analysis.model_copy(update={"reply_text": reply_text})
        except Exception:
            logger.exception("Inbox AI processing failed; using fallback reply")
            engine = "fallback"
            reply_text = build_locale_fallback_reply(final_language)
            analysis = build_default_analysis(submission.message or "", final_language, submission.source or CHAT_WIDGET_SOURCE).model_copy(
                update={"reply_text": reply_text}
            )

    logger.info(
        "Inbox analysis completed | source=%s | engine=%s | theme_slug=%s | category=%s | priority=%s",
        submission.source,
        engine,
        analysis.theme_slug,
        analysis.category,
        analysis.priority,
    )

    if analysis.reply_text != reply_text:
        analysis = analysis.model_copy(update={"reply_text": reply_text})

    try:
        with closing(get_connection()) as connection:
            thread_id = get_or_create_thread(connection, analysis)
            logger.info("Inbox SQLite save starting | thread_id=%s | database_path=%s", thread_id, settings.database_path)
            cursor = connection.execute(
                """
                INSERT INTO messages (
                    thread_id, user_name, user_email, company, source, language, category, priority, lead_score,
                    sentiment, summary, key_points_json, raw_message, reply_text, theme_label, theme_slug,
                    thread_summary, email_status, analysis_engine, created_at, sender_name, sender_email,
                    message_text, message_summary, suggested_reply, urgency_score, themes, needs_follow_up,
                    feedback_signal, feedback_reason, chat_intent, whatsapp_handoff
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    thread_id,
                    submission.name,
                    submission.email,
                    submission.company,
                    submission.source,
                    analysis.language,
                    analysis.category,
                    analysis.priority,
                    analysis.lead_score,
                    analysis.sentiment,
                    analysis.summary,
                    json.dumps(analysis.key_points),
                    submission.message,
                    reply_text,
                    analysis.theme_label,
                    analysis.theme_slug,
                    "",
                    "pending",
                    engine,
                    created_at,
                    submission.name,
                    submission.email,
                    submission.message,
                    analysis.summary,
                    reply_text,
                    PRIORITY_RANK.get(analysis.priority, 1) * 25,
                    json.dumps([analysis.theme_label]),
                    1 if analysis.priority != "low" or analysis.lead_score >= 4 else 0,
                    chat_meta["feedback_signal"],
                    chat_meta["feedback_reason"],
                    chat_meta["intent"],
                    1 if chat_meta["whatsapp_handoff"] else 0,
                ),
            )
            message_id = int(cursor.lastrowid)
            saved = True
            logger.info("Inbox SQLite save completed | message_id=%s | thread_id=%s", message_id, thread_id)

            try:
                thread = refresh_thread_rollup(connection, thread_id, analysis)
                thread_title = thread.get("thread_title")
                logger.info(
                    "Inbox thread rollup refreshed | thread_id=%s | message_count=%s | priority=%s",
                    thread_id,
                    thread.get("message_count"),
                    thread.get("priority"),
                )
            except Exception:
                logger.exception("Thread rollup refresh failed; continuing with minimal thread payload")
                thread = {
                    "id": thread_id,
                    "thread_title": analysis.thread_title,
                    "theme_slug": analysis.theme_slug,
                    "theme_label": analysis.theme_label,
                    "priority": analysis.priority,
                    "summary": analysis.summary,
                    "message_count": 1,
                    "lead_score": analysis.lead_score,
                    "updated_at": created_at,
                }
                thread_title = analysis.thread_title

            if is_chat_request and background_tasks is not None:
                related_messages = []
                email_status = "pending"
                background_tasks.add_task(
                    finalize_message_notification,
                    message_id,
                    submission,
                    analysis,
                    thread or {},
                    thread_id,
                    True,
                )
                logger.info(
                    "Inbox chat fast-path | message_id=%s | thread_id=%s | background_email=%s",
                    message_id,
                    thread_id,
                    True,
                )
            else:
                try:
                    related_messages = get_thread_messages(connection, thread_id, limit=4)
                except Exception:
                    logger.exception("Loading related thread messages failed; continuing with empty related_messages")
                    related_messages = []

                try:
                    logger.info("Inbox email trigger starting | message_id=%s | thread_id=%s", message_id, thread_id)
                    email_status = allowed_email_status(
                        send_email_notification(message_id, submission, analysis, thread or {}, related_messages)
                    )
                    logger.info("Inbox email trigger finished | message_id=%s | email_status=%s", message_id, email_status)
                except Exception:
                    logger.exception("Email failed but continuing flow")
                    email_status = "failed"

                try:
                    connection.execute("UPDATE messages SET email_status = ? WHERE id = ?", (email_status, message_id))
                    connection.commit()
                    logger.info("Inbox email status updated | message_id=%s | email_status=%s", message_id, email_status)
                except Exception:
                    logger.exception("Updating email status in SQLite failed; response will continue")
    except Exception:
        logger.exception("Inbox persistence failed; returning unsaved response payload")
        saved = False
        email_status = "skipped" if email_status == "pending" else allowed_email_status(email_status)
        related_messages = []
        thread = {
            "id": None,
            "thread_title": analysis.thread_title,
            "theme_slug": analysis.theme_slug,
            "theme_label": analysis.theme_label,
            "priority": analysis.priority,
            "summary": analysis.summary,
            "message_count": 1 if submission.message else 0,
            "lead_score": analysis.lead_score,
            "updated_at": created_at,
        }
        thread_title = analysis.thread_title

    try:
        message_payload = build_message_payload(
            submission=submission,
            analysis=analysis,
            engine=engine,
            reply_text=reply_text,
            email_status=email_status,
            saved=saved,
            message_id=message_id,
            thread_id=thread_id,
            thread_title=thread_title,
            chat_meta=chat_meta,
            created_at=created_at,
        )
        response_payload = {
            "ok": True,
            "analysis_engine": engine,
            "reply": reply_text,
            "message": message_payload,
            "thread": thread,
            "related_messages": related_messages,
        }
    except Exception:
        logger.exception("Inbox response construction failed; returning emergency minimal payload")
        emergency_reply = build_locale_fallback_reply(final_language)
        response_payload = {
            "ok": True,
            "analysis_engine": "fallback",
            "reply": emergency_reply,
            "message": {
                "reply_text": emergency_reply,
                "source": submission.source,
                "language": final_language,
                "email_status": "skipped",
                "saved": False,
            },
            "thread": thread,
            "related_messages": [],
        }

    logger.info(
        "Inbox submission completed | saved=%s | message_id=%s | thread_id=%s | whatsapp_handoff=%s | whatsapp_url=%s | response_keys=%s | message_keys=%s",
        saved,
        response_payload["message"].get("id"),
        (response_payload.get("thread") or {}).get("id") if isinstance(response_payload.get("thread"), dict) else None,
        response_payload["message"].get("whatsapp_handoff", False),
        response_payload["message"].get("whatsapp_url", ""),
        list(response_payload.keys()),
        list(response_payload["message"].keys()),
    )
    return JSONResponse(response_payload, status_code=200)


app = FastAPI(title="AI Portfolio Inbox & Insights", version="2.0.0")
# --- CORS CONFIG FROM ENV ---
allowed_origins_raw = os.getenv("ALLOWED_ORIGINS", "")
cors_debug_all_origins = os.getenv("CORS_DEBUG_ALL_ORIGINS", "").strip().lower() in {"1", "true", "yes", "on"}
allowed_origin_regex = os.getenv("ALLOWED_ORIGIN_REGEX", "").strip() or DEFAULT_CORS_ORIGIN_REGEX

configured_origins = [
    normalize_origin(origin)
    for origin in allowed_origins_raw.split(",")
    if origin.strip()
]

if cors_debug_all_origins:
    cors_allow_origins = ["*"]
    cors_allow_credentials = False
    cors_allow_origin_regex = None
else:
    cors_allow_origins = list(dict.fromkeys([*(normalize_origin(origin) for origin in DEFAULT_CORS_ORIGINS), *configured_origins]))
    cors_allow_credentials = True
    cors_allow_origin_regex = allowed_origin_regex

logger.info(
    "CORS configured | origins=%s | origin_regex=%s | credentials=%s | headers=%s",
    cors_allow_origins,
    cors_allow_origin_regex or "none",
    cors_allow_credentials,
    CORS_ALLOW_HEADERS,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_allow_origins,
    allow_origin_regex=cors_allow_origin_regex,
    allow_credentials=cors_allow_credentials,
    allow_methods=["*"],
    allow_headers=CORS_ALLOW_HEADERS,
)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def get_openai_client() -> OpenAI | None:
    if not settings.openai_api_key:
        return None
    return OpenAI(api_key=settings.openai_api_key)


def get_connection() -> sqlite3.Connection:
    connection = sqlite3.connect(settings.database_path)
    connection.row_factory = sqlite3.Row
    return connection


def existing_columns(connection: sqlite3.Connection, table_name: str) -> set[str]:
    rows = connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {row["name"] for row in rows}


def ensure_columns(connection: sqlite3.Connection, table_name: str, columns: dict[str, str]) -> None:
    current = existing_columns(connection, table_name)
    for column_name, definition in columns.items():
        if column_name not in current:
            connection.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def init_db() -> None:
    with closing(get_connection()) as connection:
        connection.executescript(
            """
            PRAGMA journal_mode=WAL;

            CREATE TABLE IF NOT EXISTS threads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                theme_slug TEXT,
                theme_label TEXT,
                thread_title TEXT,
                priority TEXT,
                summary TEXT,
                created_at TEXT,
                updated_at TEXT,
                message_count INTEGER DEFAULT 0,
                lead_score INTEGER DEFAULT 1,
                slug TEXT,
                title TEXT,
                category TEXT,
                theme_tags TEXT,
                representative_summary TEXT,
                urgency_score INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                thread_id INTEGER NOT NULL,
                user_name TEXT,
                user_email TEXT,
                company TEXT,
                source TEXT,
                language TEXT,
                category TEXT,
                priority TEXT,
                lead_score INTEGER DEFAULT 1,
                sentiment TEXT,
                summary TEXT,
                key_points_json TEXT DEFAULT '[]',
                raw_message TEXT,
                reply_text TEXT,
                theme_label TEXT,
                theme_slug TEXT,
                thread_summary TEXT,
                email_status TEXT,
                analysis_engine TEXT,
                created_at TEXT,
                sender_name TEXT,
                sender_email TEXT,
                message_text TEXT,
                message_summary TEXT,
                suggested_reply TEXT,
                urgency_score INTEGER DEFAULT 0,
                themes TEXT DEFAULT '[]',
                needs_follow_up INTEGER DEFAULT 0,
                feedback_signal TEXT DEFAULT 'none',
                feedback_reason TEXT,
                chat_intent TEXT,
                whatsapp_handoff INTEGER DEFAULT 0,
                FOREIGN KEY(thread_id) REFERENCES threads(id)
            );
            """
        )
        ensure_columns(
            connection,
            "threads",
            {
                "theme_slug": "TEXT",
                "theme_label": "TEXT",
                "thread_title": "TEXT",
                "priority": "TEXT",
                "summary": "TEXT",
                "created_at": "TEXT",
                "updated_at": "TEXT",
                "message_count": "INTEGER DEFAULT 0",
                "lead_score": "INTEGER DEFAULT 1",
                "slug": "TEXT",
                "title": "TEXT",
                "category": "TEXT",
                "theme_tags": "TEXT DEFAULT '[]'",
                "representative_summary": "TEXT",
                "urgency_score": "INTEGER DEFAULT 0",
            },
        )
        ensure_columns(
            connection,
            "messages",
            {
                "user_name": "TEXT",
                "user_email": "TEXT",
                "source": "TEXT",
                "language": "TEXT",
                "category": "TEXT",
                "priority": "TEXT",
                "lead_score": "INTEGER DEFAULT 1",
                "sentiment": "TEXT",
                "summary": "TEXT",
                "key_points_json": "TEXT DEFAULT '[]'",
                "raw_message": "TEXT",
                "reply_text": "TEXT",
                "theme_label": "TEXT",
                "theme_slug": "TEXT",
                "thread_summary": "TEXT",
                "email_status": "TEXT",
                "analysis_engine": "TEXT",
                "created_at": "TEXT",
                "feedback_signal": "TEXT DEFAULT 'none'",
                "feedback_reason": "TEXT",
                "chat_intent": "TEXT",
                "whatsapp_handoff": "INTEGER DEFAULT 0",
            },
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_messages_thread_id ON messages(thread_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_messages_created_at ON messages(created_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_messages_theme_slug ON messages(theme_slug)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_threads_theme_slug ON threads(theme_slug)")
        connection.commit()


@app.on_event("startup")
def on_startup() -> None:
    settings.database_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info(
        "SQLite startup path resolved | database_path=%s | render_disk_root=%s",
        settings.database_path,
        os.getenv("RENDER_DISK_ROOT", "").strip() or "none",
    )
    init_db()


def fallback_theme(tokens: set[str]) -> tuple[str, str]:
    for slug, label, keywords in THEME_RULES:
        if tokens & keywords:
            return slug, label
    return "general-inquiries", "General inbound inquiries"


def maybe_build_profile_grounded_analysis(submission: InboxSubmission) -> MessageAnalysis | None:
    normalized_text = normalize_match_text(submission.message or "")
    topic = detect_profile_grounding_topic(normalized_text)
    language = normalize_locale_value(submission.locale) or detect_language(submission.message or "")

    if topic:
        reply_text = build_grounded_profile_reply(language, topic)
        if not reply_text:
            return None

        topic_meta = {
            "name": ("name", "Name", "Name question"),
            "current-role": ("current-role", "Current Role", "Current role question"),
            "education": ("education", "Education", "Education question"),
            "previous-experience": ("previous-experience", "Previous Experience", "Previous experience question"),
        }
        theme_slug, theme_label, thread_title = topic_meta[topic]
        summary = clean_text(submission.message or thread_title, 220, thread_title)
        return MessageAnalysis(
            language=language,
            category="question",
            priority="medium",
            summary=summary,
            key_points=[summary],
            theme_label=theme_label,
            theme_slug=theme_slug,
            thread_title=thread_title,
            reply_text=reply_text,
            lead_score=2,
            sentiment="neutral",
        )

    tooling_key = detect_tooling_key(normalized_text)
    if tooling_key:
        reply_text = TOOLING_FACTS.get(language, TOOLING_FACTS["en"]).get(tooling_key)
        if reply_text:
            thread_title = {
                "current_tools": "Current tools question",
                "previous_tools": "Previous tools question",
                "sap": "SAP question",
                "excel": "Excel question",
                "qlik-sense": "Qlik Sense question",
                "salesforce": "Salesforce question",
                "endesa_tools": "Endesa tools question",
            }.get(tooling_key, "Tools question")
            summary = clean_text(submission.message or thread_title, 220, thread_title)
            return MessageAnalysis(
                language=language,
                category="question",
                priority="medium",
                summary=summary,
                key_points=[summary],
                theme_label="Tools",
                theme_slug="tools",
                thread_title=thread_title,
                reply_text=reply_text,
                lead_score=2,
                sentiment="neutral",
            )

    specific_experience_key = detect_specific_experience_key(normalized_text)
    if not specific_experience_key:
        return None

    experience_fact = EXPERIENCE_DETAIL_FACTS[specific_experience_key]
    reply_text = experience_fact.get(language, experience_fact["en"])
    thread_title = experience_fact["thread_title"]
    summary = clean_text(submission.message or thread_title, 220, thread_title)
    return MessageAnalysis(
        language=language,
        category="question",
        priority="medium",
        summary=summary,
        key_points=[summary],
        theme_label="Specific Experience",
        theme_slug="specific-experience",
        thread_title=thread_title,
        reply_text=reply_text,
        lead_score=2,
        sentiment="neutral",
    )


def heuristic_analysis(submission: InboxSubmission) -> tuple[MessageAnalysis, str]:
    grounded = maybe_build_profile_grounded_analysis(submission)
    if grounded is not None:
        return grounded, "profile-grounded"

    message = submission.message.strip()
    lower_text = f"{message} {submission.company or ''}".lower()
    tokens = tokenize(lower_text)
    language = detect_language(lower_text)

    question_words = {"how", "what", "can", "could", "when", "where", "why", "como", "que", "qué", "puedes", "podrias", "podrías"}
    suggestion_words = {"suggestion", "idea", "improve", "feature", "recommend", "sugerencia", "mejora", "recomiendo"}
    project_words = {"project", "hire", "freelance", "contract", "pricing", "budget", "proposal", "collaboration", "proyecto", "contratar", "presupuesto", "propuesta", "colaboracion", "colaboración"}
    bug_words = {"bug", "issue", "error", "broken", "problem", "fix", "fallo", "problema", "arreglar"}
    positive_words = {"great", "love", "excellent", "awesome", "helpful", "genial", "encanta", "excelente", "buen"}
    negative_words = {"broken", "issue", "problem", "confusing", "slow", "fallo", "problema", "confuso", "lento"}
    urgent_words = {"urgent", "asap", "today", "deadline", "immediately", "urgente", "hoy", "inmediato", "inmediatamente"}

    if tokens & bug_words:
        category = "bug report"
    elif tokens & project_words:
        category = "project inquiry"
    elif tokens & suggestion_words:
        category = "suggestion"
    elif "?" in message or tokens & question_words:
        category = "question"
    else:
        category = "general feedback"

    if tokens & negative_words:
        sentiment = "negative"
    elif tokens & positive_words:
        sentiment = "positive"
    else:
        sentiment = "neutral"

    lead_score = 2
    if category == "project inquiry":
        lead_score = 4
    if submission.company:
        lead_score += 1
    if tokens & {"enterprise", "company", "team", "teams", "empresa", "equipo"}:
        lead_score += 1
    lead_score = max(1, min(5, lead_score))

    priority = "low"
    if category == "bug report" and tokens & urgent_words:
        priority = "high"
    elif category == "project inquiry" and (lead_score >= 4 or tokens & urgent_words):
        priority = "high"
    elif category in {"suggestion", "question", "bug report"} or lead_score >= 3:
        priority = "medium"

    theme_slug, theme_label = fallback_theme(tokens)
    thread_title = f"{theme_label} thread"
    sentence_parts = [part.strip() for part in re.split(r"(?<=[.!?])\s+", message) if part.strip()]
    summary = (sentence_parts[0] if sentence_parts else message)[:220]
    key_points = normalize_list(sentence_parts[:3] or [summary], limit=5)

    if language == "es":
        reply_text = "Gracias por escribir. Tu mensaje se ha registrado correctamente. Si hace falta, puedo ampliar experiencia, proyectos o encaje profesional desde un perfil centrado en operaciones, reporting, soporte a procesos y uso practico de IA."
    else:
        reply_text = "Thanks for reaching out. Your message has been logged successfully. If helpful, I can expand on experience, projects, or professional fit from a profile centered on operations, reporting, process support, and practical AI use."

    return (
        MessageAnalysis(
            language=language,
            category=category,
            priority=priority,
            summary=summary,
            key_points=key_points,
            theme_label=theme_label,
            theme_slug=theme_slug,
            thread_title=thread_title,
            reply_text=reply_text,
            lead_score=lead_score,
            sentiment=sentiment,
        ),
        "heuristic",
    )


def detect_chat_intent(normalized_text: str) -> str:
    recruiter_terms = {
        "recruiter", "recruiting", "hiring", "hire", "hiring manager", "talent", "talent acquisition",
        "role", "roles", "position", "positions", "job", "opening", "opportunity", "interview", "team",
        "fit", "candidate", "cv", "resume", "background", "profile for", "perfil", "encaje", "puesto",
        "vacante", "reclutador", "seleccion", "seleccionador", "manager", "manager hiring",
    }
    client_terms = {
        "client", "project", "collaboration", "consulting", "consultant", "consultancy", "freelance",
        "support", "workflow", "process", "process improvement", "operations support", "dashboard",
        "reporting", "reporting help", "operations", "automation", "ai", "partner", "proposal",
        "quote", "scope", "need help", "help with", "improve", "improvement", "efficiency",
        "cliente", "proyecto", "colaboracion", "consultoria", "proceso", "mejora de procesos",
        "reporte", "reporting", "dashboard", "operaciones", "soporte operativo", "automatizacion",
        "propuesta", "presupuesto", "colaborador",
    }
    if any(term in normalized_text for term in recruiter_terms):
        return "recruiter"
    if any(term in normalized_text for term in client_terms):
        return "client"
    return "general"


def detect_summary_request(normalized_text: str) -> bool:
    summary_patterns = [
        "summary", "summarize", "summarise", "short version", "short overview", "quick summary",
        "high level overview", "key points", "overview", "resumen", "resumelo", "hazme un resumen",
        "dime lo principal", "resumen rapido", "resumen en 3 puntos", "lo mas importante",
    ]
    return any(pattern in normalized_text for pattern in summary_patterns)


def detect_profile_grounding_topic(normalized_text: str) -> str | None:
    name_terms = {
        "como te llamas", "como se llama", "cual es tu nombre", "quien eres", "who are you", "what is your name",
        "your name", "whats your name", "what's your name", "who is this",
    }
    current_role_terms = {
        "where does", "works at", "work at", "current company", "current employer", "current job",
        "where does carlos work", "what company does he work for", "empresa actual", "donde trabaja",
        "donde trabaja actualmente", "donde trabaja ahora", "trabaja actualmente", "en que empresa", "current role",
        "trabajo actual", "en que trabajas", "en que trabaja", "y en que trabaja", "a que te dedicas",
        "a que se dedica", "en que trabajas ahora", "en que trabaja ahora", "rpc", "retail performance company",
    }
    education_terms = {
        "education", "academic background", "studies", "study", "studied", "what has he studied",
        "what does he study", "formacion", "estudios", "que estudia", "que ha estudiado",
        "que estudiaste", "que has estudiado", "que formacion tienes", "formacion academica", "academic studies",
    }
    previous_experience_terms = {
        "previous experience", "prior experience", "worked before", "where has he worked before",
        "what has he worked on before", "career history", "experiencia previa", "ha trabajado antes",
        "donde ha trabajado antes", "donde trabajaste antes", "en que ha trabajado", "en que trabajabas antes",
        "que experiencia tienes", "que hacias antes", "que hacias antes de rpc", "antes de rpc",
        "antes de trabajar en rpc", "antes de rpc?", "trabajos anteriores", "trayectoria previa",
    }
    if "antes" in normalized_text and "rpc" in normalized_text:
        return "previous-experience"
    if any(term in normalized_text for term in name_terms):
        return "name"
    if any(term in normalized_text for term in previous_experience_terms):
        return "previous-experience"
    if any(term in normalized_text for term in education_terms):
        return "education"
    if any(term in normalized_text for term in current_role_terms):
        return "current-role"
    return None


def detect_specific_experience_key(normalized_text: str) -> str | None:
    for key, item in EXPERIENCE_DETAIL_FACTS.items():
        if any(term in normalized_text for term in item["terms"]):
            return key
    return None


def detect_tooling_key(normalized_text: str) -> str | None:
    current_markers = {
        "herramientas utilizas", "herramientas usas", "con que herramientas trabajas", "con que herramientas usas",
        "software utilizas", "herramientas utilizas actualmente", "herramientas usas actualmente",
        "herramientas usas en tu trabajo actual", "que herramientas utilizas", "que herramientas usas",
        "what tools do you use", "which tools do you use", "what software do you use", "current tools",
    }
    previous_markers = {
        "herramientas has utilizado", "herramientas usabas", "otros proyectos", "experiencias previas",
        "herramientas de experiencias anteriores", "what tools have you used", "previous tools",
        "tools in previous experience",
    }

    if detect_whatsapp_request(normalized_text) or detect_direct_contact_request(normalized_text):
        return None
    if "salesforce" in normalized_text:
        return "salesforce"
    if "qlik sense" in normalized_text:
        return "qlik-sense"
    sap_context_terms = {
        "usas sap", "trabajas con sap", "experiencia con sap", "herramienta sap", "software sap",
        "sap en tu trabajo", "sap en rpc", "use sap", "work with sap", "experience with sap", "sap tool",
    }
    if contains_standalone_term(normalized_text, "sap") and (
        any(term in normalized_text for term in sap_context_terms) or "sap?" in normalized_text or normalized_text == "sap"
    ):
        return "sap"
    if "excel" in normalized_text:
        return "excel"
    if "endesa" in normalized_text and ("herramient" in normalized_text or "tool" in normalized_text or "software" in normalized_text):
        return "endesa_tools"
    if any(term in normalized_text for term in previous_markers):
        return "previous_tools"
    if any(term in normalized_text for term in current_markers):
        return "current_tools"
    return None


def detect_chat_topic(normalized_text: str) -> str:
    if detect_summary_request(normalized_text):
        return "summary"
    grounding_topic = detect_profile_grounding_topic(normalized_text)
    if grounding_topic:
        return grounding_topic
    if detect_whatsapp_request(normalized_text):
        return "whatsapp"
    if any(term in normalized_text for term in {"contact", "email", "reach", "connect", "linkedin", "contactar", "correo", "hablar"}):
        return "contact"
    if any(term in normalized_text for term in {"project", "projects", "portfolio", "case study", "proyecto", "proyectos", "examples of his work", "best work", "main projects", "principales proyectos", "mejores proyectos"}):
        return "projects"
    if any(term in normalized_text for term in {"experience", "background", "career", "trayectoria", "experiencia"}):
        return "experience"
    if any(term in normalized_text for term in {"fit", "role", "roles", "encaje", "puesto", "perfil"}):
        return "fit"
    if any(term in normalized_text for term in {"ai", "automation", "workflow", "productivity", "ia", "automatizacion", "productividad"}):
        return "practical-ai"
    if any(term in normalized_text for term in {"about", "who is", "what does", "profile", "sobre", "quien es", "que hace"}):
        return "profile"
    return "general"


def detect_contact_readiness(normalized_text: str) -> bool:
    contact_terms = {
        "contact", "reach", "email", "call", "meeting", "book", "connect", "talk", "speak", "follow up",
        "linkedin", "whatsapp", "correo", "contactar", "llamar", "reunion", "agendar", "hablar", "escribir",
        "ponernos en contacto", "seguir hablando",
    }
    return detect_whatsapp_request(normalized_text) or any(term in normalized_text for term in contact_terms)


def needs_guided_next_step(normalized_text: str, topic: str) -> bool:
    broad_terms = {
        "tell me about", "what does he do", "how can he help", "what kind of profile", "overview",
        "more about", "cuentame", "hablame", "que hace", "que perfil", "como puede ayudar",
        "mas informacion", "mas sobre", "dime sobre",
    }
    return topic == "general" and any(term in normalized_text for term in broad_terms)


def detect_whatsapp_request(normalized_text: str) -> bool:
    whatsapp_terms = {
        "whatsapp", "whats app", "wa.me", "escribirte por whatsapp", "hablar por whatsapp",
        "contact on whatsapp", "talk on whatsapp", "message on whatsapp", "pasame tu whatsapp",
        "pasa tu whatsapp", "quiero escribirte por whatsapp", "quiero hablar por whatsapp",
        "quiero contactar por whatsapp", "prefiero whatsapp", "do you have whatsapp", "wasap",
        "guasap", "watsapp", "whatsap", "hablar por wasap", "conectar por wasap",
        "contactar por wasap", "contacto por whatsapp", "abrir whatsapp", "escribeme por wasap",
        "escribirme por wasap", "hablar directamente por whatsapp", "conversacion por wasap",
        "quiero una conversacion por wasap", "quiero conectarme por wasap", "quiero hablar con el por wasap",
    }
    return any(term in normalized_text for term in whatsapp_terms)


def detect_direct_contact_request(normalized_text: str) -> bool:
    direct_terms = {
        "human", "person", "directly", "direct conversation", "direct contact", "speak with carlos",
        "talk to carlos", "speak to a person", "real person", "persona", "humano", "directamente",
        "hablar contigo", "hablar con carlos", "hablar con una persona", "contacto directo",
        "message directly", "write directly", "escribirte directamente", "quiero escribirte directamente",
        "via mas directa", "vía mas directa", "una persona", "quiero contactar", "quiero hablar",
        "quiero contactar con carlos", "quiero hablar con carlos", "quiero una conversacion directa",
        "quiero una conversación directa", "quiero escribirle", "quiero hablar sobre un proyecto",
        "quiero colaborar", "quiero contactar por whatsapp",
    }
    return any(term in normalized_text for term in direct_terms)


def detect_chat_frustration(normalized_text: str) -> bool:
    frustration_terms = {
        "not helping", "this is not helping", "not useful", "not what i asked", "you dont understand",
        "you are not understanding", "still not", "this doesnt help", "that is not it", "frustrating",
        "no me ayuda", "no me estas ayudando", "no me estás ayudando", "no me sirve", "no es eso",
        "sigues sin entender", "no entiendes", "esto no ayuda", "esto no me sirve", "mejor hablar con alguien",
    }
    return any(term in normalized_text for term in frustration_terms)


def detect_conversation_close_intent(normalized_text: str) -> bool:
    close_terms = {
        "adios", "hasta luego", "hasta pronto", "gracias ya esta", "gracias ya esta.", "ya esta",
        "ya esta gracias", "eso es todo", "no quiero seguir", "no quiero hablar mas", "no quiero hablar mas contigo",
        "vale gracias", "perfecto nada mas", "nada mas", "no necesito mas", "gracias por todo",
        "goodbye", "bye", "thanks thats all", "thanks that's all", "thats all", "that's all",
        "i dont want to continue", "i do not want to continue", "i dont want to keep talking",
        "no more questions", "all good thanks", "all good, thanks", "thats enough", "that's enough",
    }
    return any(term in normalized_text for term in close_terms)


def detect_feedback_signal(normalized_text: str) -> tuple[str, str | None]:
    negative_patterns = {
        "wrong_answer": {
            "no es eso", "eso no responde", "sigues sin entender", "that is not what i asked",
            "thats not what i asked", "you still dont understand", "you still do not understand",
        },
        "too_vague": {
            "no me sirve", "too vague", "too generic", "be more specific", "mas concreto", "más concreto",
            "muy generico", "muy genérico", "demasiado vago",
        },
        "too_long": {
            "too long", "demasiado largo", "mas corto", "más corto", "resumelo mas", "resúmelo más",
        },
        "not_helpful": {
            "this is not helping", "not helping", "esto no ayuda", "no me ayuda", "eso no ayuda",
        },
    }
    positive_patterns = {
        "helpful": {"gracias eso ayuda", "gracias, eso ayuda", "thanks that helps", "thanks, that helps", "that helps"},
        "exact": {"perfecto", "eso era", "exactly", "that works", "great", "perfect", "justo eso"},
    }

    for reason, patterns in negative_patterns.items():
        if any(pattern in normalized_text for pattern in patterns):
            return "negative", reason
    for reason, patterns in positive_patterns.items():
        if any(pattern in normalized_text for pattern in patterns):
            return "positive", reason
    return "none", None


def is_scope_gap_question(normalized_text: str, topic: str) -> bool:
    if topic != "general":
        return False

    question_terms = {
        "what", "which", "where", "when", "why", "how", "who",
        "que", "que ", "cual", "cuales", "donde", "cuando", "por que", "quien", "como",
        "tell me", "can you explain", "podrias explicar", "puedes explicar",
    }
    return "?" in normalized_text or any(term in normalized_text for term in question_terms)


def detect_recent_feedback(messages: list[dict[str, str]] | None) -> tuple[str, str | None]:
    if not messages:
        return "none", None
    recent_user_messages = [
        normalize_match_text(item.get("content", ""))
        for item in reversed(messages[-6:])
        if item.get("role") == "user" and item.get("content")
    ]
    for message in recent_user_messages[:2]:
        signal, reason = detect_feedback_signal(message)
        if signal != "none":
            return signal, reason
    return "none", None


def detect_stuck_conversation(messages: list[dict[str, str]] | None) -> bool:
    if not messages:
        return False

    recent_user_messages = [
        normalize_match_text(item.get("content", ""))
        for item in reversed(messages[-8:])
        if item.get("role") == "user" and item.get("content")
    ]
    recent_user_messages = [message for message in recent_user_messages if message]
    if len(recent_user_messages) < 3:
        return False

    current_tokens = tokenize(recent_user_messages[0])
    if len(current_tokens) < 2:
        return False

    overlaps = 0
    for previous in recent_user_messages[1:3]:
        previous_tokens = tokenize(previous)
        if len(current_tokens & previous_tokens) >= 2:
            overlaps += 1
    return overlaps >= 2


def normalize_reply_signature(text: str) -> str:
    cleaned = normalize_match_text(re.sub(r"https?://\S+", "", text or ""))
    return cleaned.strip()


def should_use_repetition_fallback(candidate_reply: str, messages: list[dict[str, str]] | None, topic: str) -> bool:
    if not candidate_reply or not messages:
        return False
    if topic in {"current-role", "education", "previous-experience", "specific-experience", "tools", "name"}:
        return False

    candidate_signature = normalize_reply_signature(candidate_reply)
    if not candidate_signature:
        return False

    recent_assistant_messages = [
        normalize_reply_signature(item.get("content", ""))
        for item in reversed(messages[-6:])
        if item.get("role") == "assistant" and item.get("content")
    ]
    recent_assistant_messages = [item for item in recent_assistant_messages if item]
    if not recent_assistant_messages:
        return False

    if candidate_signature == recent_assistant_messages[0]:
        return True

    repeated_count = sum(1 for item in recent_assistant_messages[:2] if item == candidate_signature)
    return repeated_count >= 2


def build_whatsapp_cta(language: str, reason: str) -> str:
    whatsapp_url = build_whatsapp_url(language)
    if language == "es":
        if reason == "explicit":
            return f"Claro, puedes escribirme directamente por WhatsApp. {whatsapp_url} Si quieres, dime antes si buscas una oportunidad profesional, una colaboracion o informacion sobre un proyecto."
        if reason in {"frustration", "stuck", "direct"}:
            return f"Si prefieres una conversacion mas directa, puedes escribirme por WhatsApp. {whatsapp_url} Si te va bien, cuentame brevemente que tipo de conversacion buscas."
        return f"Si quieres hablar directamente sobre la oportunidad o colaboracion, puedes escribirme por WhatsApp. {whatsapp_url} Si quieres, dime antes si esto es por un rol, una colaboracion o un proyecto."

    if reason == "explicit":
        return f"Sure, you can contact Carlos directly on WhatsApp. {whatsapp_url} If you want, tell me briefly whether this is about a role, a collaboration, or a project."
    if reason in {"frustration", "stuck", "direct"}:
        return f"If you prefer a more direct conversation, you can reach Carlos on WhatsApp. {whatsapp_url} If helpful, let me know what kind of conversation you're looking for."
    return f"If you want to discuss the opportunity or collaboration directly, you can contact Carlos on WhatsApp. {whatsapp_url} If helpful, tell me whether this is about a role, a collaboration, or a project."


def resolve_whatsapp_offer(
    language: str,
    intent: str,
    topic: str,
    normalized_text: str,
    messages: list[dict[str, str]] | None,
) -> tuple[bool, str | None]:
    whatsapp_requested = detect_whatsapp_request(normalized_text)
    if whatsapp_requested:
        return True, "explicit"

    if detect_chat_frustration(normalized_text):
        return True, "frustration"

    if detect_direct_contact_request(normalized_text):
        return True, "direct"

    if detect_stuck_conversation(messages):
        return True, "stuck"

    if topic == "contact" and intent in {"recruiter", "client"}:
        return True, "professional"

    return False, None


def build_feedback_recovery_reply(language: str, intent: str, whatsapp_offer: bool, whatsapp_reason: str | None) -> str:
    whatsapp_tail = build_whatsapp_cta(language, whatsapp_reason or "direct") if whatsapp_offer and whatsapp_reason else ""
    if language == "es":
        base = "Entiendo. Voy a ir mas directo. Puedo enfocarlo en experiencia, proyectos o encaje profesional."
        return f"{base} {whatsapp_tail}".strip()
    base = "Understood. I'll keep it more direct. I can focus on experience, projects, or role fit."
    return f"{base} {whatsapp_tail}".strip()


def build_scope_gap_reply(language: str, whatsapp_offer: bool, whatsapp_reason: str | None) -> str:
    whatsapp_url = build_whatsapp_url(language)
    if language == "es":
        return f"No tengo esa informacion con suficiente precision. Puedes consultarmelo escribiendo en el formulario o a traves de WhatsApp. {whatsapp_url}"
    return f"I do not have that information with enough precision. You can ask through the contact form or on WhatsApp. {whatsapp_url}"


def build_conversation_close_reply(language: str) -> str:
    whatsapp_url = build_whatsapp_url(language)
    if language == "es":
        return f"Encantado de haberte ayudado. Para cualquier otra duda puedes escribirme a traves del formulario o abrir conversacion por WhatsApp. {whatsapp_url}"
    return f"Glad I could help. If you need anything else, you can reach out through the contact form or open a WhatsApp conversation. {whatsapp_url}"


def build_chat_cta(language: str, intent: str, topic: str, contact_ready: bool, whatsapp_offer: bool, whatsapp_reason: str | None) -> str:
    if language == "es":
        if contact_ready or topic == "contact":
            return f"Puedes escribirme a traves del formulario del portfolio o, si prefieres una conversacion mas directa, abrir conversacion por WhatsApp. {build_whatsapp_url(language)}"
        if whatsapp_offer and whatsapp_reason:
            return build_whatsapp_cta(language, whatsapp_reason)
        if intent == "recruiter":
            return "Si quieres, puedo resumir su encaje para un rol, destacar la experiencia mas relevante o centrarme en proyectos."
        if intent == "client":
            return "Si te sirve, puedo enfocarlo a procesos, reporting, workflows digitales o IA practica."
        return "Si quieres, puedo resumirlo en corto, enfocarlo a experiencia o senalar los proyectos mas relevantes."

    if contact_ready or topic == "contact":
        return f"You can write through the portfolio contact form or, if you prefer a more direct conversation, open a WhatsApp chat. {build_whatsapp_url(language)}"
    if whatsapp_offer and whatsapp_reason:
        return build_whatsapp_cta(language, whatsapp_reason)
    if intent == "recruiter":
        return "If useful, I can summarise his fit for a role, highlight the most relevant experience, or focus on projects."
    if intent == "client":
        return "If helpful, I can frame his value around processes, reporting, digital workflows, or practical AI."
    return "If you want, I can keep it short, focus on experience, or point you to the most relevant projects."


def build_guided_options(language: str, intent: str) -> str:
    if language == "es":
        if intent == "recruiter":
            return "Puedo ayudarte de tres formas: resumen de perfil, encaje para un rol o experiencia mas relevante."
        if intent == "client":
            return "Puedo ayudarte de tres formas: resumen del perfil, donde puede aportar en procesos/reporting o siguiente paso de contacto."
        return "Puedo ayudarte de tres formas: resumen del perfil, experiencia principal o proyectos mas relevantes."
    if intent == "recruiter":
        return "I can help in three ways: a profile summary, fit for a role, or the most relevant experience."
    if intent == "client":
        return "I can help in three ways: a profile summary, where he can add value in processes/reporting, or how to get in touch."
    return "I can help in three ways: a profile summary, key experience, or the most relevant projects."


def build_projects_reply(language: str) -> str:
    entries = PROJECT_SPOTLIGHTS.get(language, PROJECT_SPOTLIGHTS["en"])
    if language == "es":
        body = "\n".join(
            f"- {item['title']} - {item['summary']} {item['value']}"
            for item in entries[:3]
        )
        return (
            "Estos son los proyectos mas relevantes:\n"
            + body
            + "\n\nSi quieres, puedo resumir uno de ellos con mas detalle o decirte cual encaja mejor con operaciones, datos o IA practica."
        )

    body = "\n".join(
        f"- {item['title']} - {item['summary']} {item['value']}"
        for item in entries[:3]
    )
    return (
        "These are the most relevant projects:\n"
        + body
        + "\n\nIf you want, I can expand on one of them or tell you which one is most relevant for operations, data, or practical AI."
    )


def build_education_reply(language: str) -> str:
    facts = PORTFOLIO_ROLE_FACTS.get(language, PORTFOLIO_ROLE_FACTS["en"])["education"]
    if language == "es":
        body = "\n".join(f"- {item}" for item in facts)
        return (
            "En formacion, el portfolio muestra esto:\n"
            + body
            + "\n\nSi quieres, puedo resumirlo en una version mas breve o relacionarlo con su perfil profesional actual."
        )
    body = "\n".join(f"- {item}" for item in facts)
    return (
        "In education, the portfolio shows this:\n"
        + body
        + "\n\nIf you want, I can keep it shorter or connect it to his current professional profile."
    )


def build_previous_experience_reply(language: str) -> str:
    facts = PORTFOLIO_ROLE_FACTS.get(language, PORTFOLIO_ROLE_FACTS["en"])["previous_experience"]
    if language == "es":
        body = "\n".join(f"- {item}" for item in facts)
        return (
            "En experiencia previa, el portfolio recoge estas etapas:\n"
            + body
            + "\n\nSi te interesa, tambien puedo resumir como esa trayectoria conecta con operaciones, reporting y soporte a procesos."
        )
    body = "\n".join(f"- {item}" for item in facts)
    return (
        "For previous experience, the portfolio includes these roles:\n"
        + body
        + "\n\nIf useful, I can also summarise how that background connects with operations, reporting, and process support."
    )


def build_current_role_reply(language: str) -> str:
    base = PORTFOLIO_ROLE_FACTS.get(language, PORTFOLIO_ROLE_FACTS["en"])["current_role"]
    if language == "es":
        return (
            base
            + "\n\nSi quieres, tambien puedo resumir sus funciones principales o explicarte como encaja ese rol con su perfil general."
        )
    return (
        base
        + "\n\nIf you want, I can also summarise his main responsibilities or explain how that role fits his broader profile."
    )


def build_name_reply(language: str) -> str:
    if language == "es":
        return "Mi nombre es Carlos San Miguel."
    return "My name is Carlos San Miguel."


def build_grounded_profile_reply(language: str, topic: str) -> str | None:
    if topic == "name":
        return build_name_reply(language)
    if topic == "current-role":
        return build_current_role_reply(language)
    if topic == "education":
        return build_education_reply(language)
    if topic == "previous-experience":
        return build_previous_experience_reply(language)
    return None


def build_static_chat_reply(
    language: str,
    topic: str,
    intent: str,
    contact_ready: bool,
    guided_mode: bool,
    whatsapp_offer: bool,
    whatsapp_reason: str | None,
) -> str | None:
    tail = build_chat_cta(language, intent, topic, contact_ready, whatsapp_offer, whatsapp_reason)

    replies = {
        "en": {
            "whatsapp": build_whatsapp_cta("en", whatsapp_reason or "explicit"),
            "summary": "Quick summary:\n- Carlos is an operations, data, and digital support professional with a clear corporate profile.\n- His experience is centered on process support, reporting, SAP-related workflows, coordination, incidents, and structured follow-up.\n- He also uses digital tools and practical AI to improve productivity, information handling, and workflow efficiency.",
            "current-role": build_current_role_reply("en") + "\n\n" + tail,
            "education": build_education_reply("en") + "\n\n" + tail,
            "previous-experience": build_previous_experience_reply("en") + "\n\n" + tail,
            "profile": "Carlos' profile combines operations, data visibility, digital workflows, and practical AI use. He currently works at The Retail Performance Company (RPC), where his focus is on process coordination, reporting, SAP-related support, and helping teams work with more structure and clarity.\n\n" + tail,
            "experience": "Carlos currently works at The Retail Performance Company (RPC) since November 2025. His recent experience is centered on operations, Purchasing and Aftersales support, SAP-related workflows, reporting follow-up, incident handling, and coordination in a BMW-related corporate environment. His broader background also includes back-office operations, documentation validation, public procurement support, banking operations, and technical support operations.\n\n" + tail,
            "projects": build_projects_reply("en"),
            "fit": "Carlos is a strong fit for roles that combine operations, reporting, process support, procurement support, digital coordination, and practical AI use. He is especially relevant for teams that need structured follow-up, business support, KPI visibility, and someone comfortable working across tools, workflows, and operational contexts.\n\n" + tail,
            "practical-ai": "His use of AI is practical and business-oriented. The focus is on writing support, information organization, summarization, and workflow efficiency in everyday workflows.\n\n" + tail,
            "contact": ("You can use the contact option in the portfolio if you want to discuss a role, collaboration, or project context. " + tail) if whatsapp_offer else "You can use the contact option in the portfolio if you want to discuss a role, collaboration, or project context. If helpful first, I can also summarise his experience or highlight the most relevant projects before you reach out.",
            "general": "Carlos' portfolio presents a professional profile centered on operations, reporting, process support, digital workflows, and practical AI. The strongest themes are corporate business support, data visibility, coordination, and useful digital initiatives.\n\n" + (build_guided_options(language, intent) if guided_mode else tail),
        },
        "es": {
            "whatsapp": build_whatsapp_cta("es", whatsapp_reason or "explicit"),
            "summary": "Resumen rapido:\n- Carlos es un profesional de operaciones, datos y soporte digital con perfil corporativo claro.\n- Su experiencia se centra en soporte a procesos, reporting, workflows relacionados con SAP, coordinacion, incidencias y seguimiento estructurado.\n- Tambien utiliza herramientas digitales e IA practica para mejorar productividad, gestion de la informacion y eficiencia de workflows.",
            "current-role": build_current_role_reply("es") + "\n\n" + tail,
            "education": build_education_reply("es") + "\n\n" + tail,
            "previous-experience": build_previous_experience_reply("es") + "\n\n" + tail,
            "profile": "El perfil de Carlos combina operaciones, visibilidad de datos, workflows digitales e IA practica. Actualmente trabaja en The Retail Performance Company (RPC), donde sus puntos mas fuertes estan en coordinacion de procesos, reporting, soporte relacionado con SAP y apoyo a equipos con mas estructura y claridad.\n\n" + tail,
            "experience": "Carlos trabaja actualmente en The Retail Performance Company (RPC) desde noviembre de 2025. Su experiencia reciente se centra en operaciones, soporte a Purchasing y Aftersales, workflows relacionados con SAP, seguimiento de reporting, gestion de incidencias y coordinacion en un entorno corporativo vinculado a BMW. Su trayectoria tambien incluye back office, validacion documental, contratacion publica, operaciones bancarias y soporte tecnico-operativo.\n\n" + tail,
            "projects": build_projects_reply("es"),
            "fit": "Carlos encaja bien en roles que combinan operaciones, reporting, soporte a procesos, soporte a compras, coordinacion digital y uso practico de IA. Es especialmente relevante para equipos que necesitan seguimiento estructurado, soporte de negocio, visibilidad KPI y alguien comodo trabajando entre herramientas, workflows y contextos operativos.\n\n" + tail,
            "practical-ai": "Su uso de IA es practico y orientado a negocio. El foco esta en apoyo a redaccion, organizacion de informacion, resumenes y eficiencia de workflows del dia a dia.\n\n" + tail,
            "contact": ("Puedes usar la opcion de contacto del portfolio si quieres hablar sobre un rol, una colaboracion o un contexto de proyecto. " + tail) if whatsapp_offer else "Puedes usar la opcion de contacto del portfolio si quieres hablar sobre un rol, una colaboracion o un contexto de proyecto. Si te ayuda antes, tambien puedo resumir su experiencia o destacar los proyectos mas relevantes.",
            "general": "El portfolio de Carlos presenta un perfil profesional centrado en operaciones, reporting, soporte a procesos, workflows digitales e IA practica. Los temas mas fuertes son soporte corporativo a negocio, visibilidad de datos, coordinacion e iniciativas digitales utiles.\n\n" + (build_guided_options(language, intent) if guided_mode else tail),
        },
    }
    return replies.get(language, replies["en"]).get(topic)


def fallback_chat_reply(language: str, intent: str, contact_ready: bool, whatsapp_offer: bool, whatsapp_reason: str | None) -> str:
    return build_scope_gap_reply(language, whatsapp_offer, whatsapp_reason)


def build_chat_system_prompt(
    language: str,
    intent: str,
    topic: str,
    contact_ready: bool,
    guided_mode: bool,
    whatsapp_offer: bool,
    whatsapp_reason: str | None,
    feedback_signal: str,
    feedback_reason: str | None,
) -> str:
    brevity_rule = (
        "Responde en un maximo de 3 bullets cortos."
        if topic == "summary" and language == "es"
        else "Answer in a maximum of 3 short bullet points."
        if topic == "summary"
        else "Manten la respuesta breve, clara y facil de escanear."
        if language == "es"
        else "Keep the reply concise, high-signal, and easy to scan."
    )
    whatsapp_rule = (
        f"Solo ofrece WhatsApp cuando corresponda. Si debes derivar a WhatsApp, usa exactamente este numero {WHATSAPP_NUMBER} y este enlace {WHATSAPP_LINK}. Motivo actual: {whatsapp_reason or 'none'}."
        if language == "es"
        else f"Offer WhatsApp only when it is clearly appropriate. If you should offer WhatsApp, use exactly this number {WHATSAPP_NUMBER} and this link {WHATSAPP_LINK}. Current reason: {whatsapp_reason or 'none'}."
    )
    feedback_rule = (
        f"Feedback actual del usuario: {feedback_signal} ({feedback_reason or 'none'}). Si es negativo, responde mas corto, mas concreto y ofrece 2 o 3 opciones claras. Si es positivo, manten el nivel de detalle."
        if language == "es"
        else f"Current user feedback: {feedback_signal} ({feedback_reason or 'none'}). If it is negative, be shorter, more concrete, and offer 2 or 3 clear options. If it is positive, keep the current level of detail."
    )

    if language == "es":
        return f"""
Responde SOLO en espanol.
No respondas en ingles salvo que el usuario pida claramente cambiar a ingles.
Manten el idioma de la conversacion salvo cambio explicito del usuario.
Los resumenes, CTAs y siguientes pasos deben salir tambien en espanol.

Eres el asistente del portfolio de Carlos San Miguel.
Debes sonar profesional, claro y recruiter-friendly.
Situa a Carlos en operaciones, datos, workflows digitales e IA practica.
No inventes experiencia ni exageres seniority.
Habla siempre en positivo y centrado en valor real. No definas el perfil por negacion.
Adapta la respuesta a esta intencion: {intent}.
{brevity_rule}
Piensa como un asistente profesional de portfolio, no como un bot comercial.
Si la intencion es recruiter o hiring, conecta la respuesta con encaje, experiencia relevante y un siguiente paso de contacto razonable.
Si la intencion es cliente o colaboracion, conecta la respuesta con procesos, reporting, workflows digitales e IA practica.
Si la pregunta es amplia ({guided_mode}), sintetiza primero y luego ofrece 2 o 3 caminos claros.
Si la intencion de contacto es explicita ({contact_ready}), orienta el siguiente paso de forma practica y sin presion.
 {whatsapp_rule}
{feedback_rule}
Termina con un siguiente paso breve y natural solo cuando realmente ayude.
""".strip()

    return f"""
Respond ONLY in English.
Do not answer in Spanish unless the user clearly asks to switch to Spanish.
Maintain the conversation language unless the user explicitly changes it.
Summaries, CTAs, and next steps must also stay in English.

You are the portfolio assistant for Carlos San Miguel.
Sound professional, clear, and recruiter-friendly.
Position Carlos around operations, data, digital workflows, and practical AI.
Do not invent experience or exaggerate seniority.
Always use positive framing centered on what Carlos contributes. Do not define the profile through negation.
Adapt to this intent: {intent}.
{brevity_rule}
Think like a professional portfolio assistant, not a sales bot.
For recruiter or hiring intent, connect the answer to role fit, relevant experience, and a sensible contact path.
For client or collaboration intent, connect the answer to process support, reporting, digital workflows, and practical AI use.
If the question is broad ({guided_mode}), synthesize first and then offer 2 or 3 clear directions.
If contact intent is explicit ({contact_ready}), make the next step practical and low-pressure.
{whatsapp_rule}
{feedback_rule}
End with one brief, natural next step only when it genuinely helps.
""".strip()


def translate_chat_labels(language: str) -> dict[str, str]:
    if language == "es":
        return {
            "facts": "Hechos relevantes",
            "conversation": "Conversacion reciente",
        }
    return {
        "facts": "Relevant facts",
        "conversation": "Recent conversation",
    }


def openai_chat_reply(
    submission: InboxSubmission,
    analysis: MessageAnalysis,
    language: str,
    intent: str,
    topic: str,
    contact_ready: bool,
    guided_mode: bool,
    whatsapp_offer: bool,
    whatsapp_reason: str | None,
    feedback_signal: str,
    feedback_reason: str | None,
) -> tuple[str | None, str]:
    client = get_openai_client()
    if client is None:
        return None, "chat-heuristic"

    history = submission.messages or []
    filtered_history = [
        {
            "role": item.get("role", "user") if item.get("role") in {"assistant", "user"} else "user",
            "content": item.get("content", "").strip(),
        }
        for item in history[-6:]
        if item.get("content")
    ]

    if not filtered_history:
        filtered_history = [{"role": "user", "content": submission.message}]

    facts = "\n".join(f"- {fact}" for fact in CHAT_PROFILE_FACTS.get(language, CHAT_PROFILE_FACTS["en"]))
    history_text = "\n".join(f"{item['role']}: {item['content']}" for item in filtered_history)
    system_prompt = build_chat_system_prompt(
        language,
        intent,
        topic,
        contact_ready,
        guided_mode,
        whatsapp_offer,
        whatsapp_reason,
        feedback_signal,
        feedback_reason,
    )
    labels = translate_chat_labels(language)
    context_block = f"{labels['facts']}:\n{facts}\n\n{labels['conversation']}:\n{history_text}"
    try:
        response = client.responses.create(
            model=settings.openai_model,
            input=[
                {"role": "system", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": [{"type": "input_text", "text": context_block}]},
            ],
            max_output_tokens=220,
        )
        reply = response.output_text.strip()
        engine = "chat-openai"
        if whatsapp_offer and whatsapp_reason:
            engine = f"{engine}-whatsapp-{whatsapp_reason}"
        return (reply or None), engine
    except Exception as exc:
        logger.info("Chat reply fallback used after OpenAI error: %s", exc)
        return None, "chat-heuristic"


def generate_chat_reply(submission: InboxSubmission, analysis: MessageAnalysis, language: str) -> tuple[str, str, dict[str, Any]]:
    normalized = normalize_match_text(submission.message)
    intent = detect_chat_intent(normalized)
    topic = detect_chat_topic(normalized)
    contact_ready = detect_contact_readiness(normalized)
    guided_mode = needs_guided_next_step(normalized, topic)
    whatsapp_offer, whatsapp_reason = resolve_whatsapp_offer(language, intent, topic, normalized, submission.messages)
    if contact_ready or topic == "contact":
        whatsapp_offer = True
        whatsapp_reason = whatsapp_reason or "direct"
    feedback_signal, feedback_reason = detect_feedback_signal(normalized)
    if feedback_signal == "none":
        feedback_signal, feedback_reason = detect_recent_feedback(submission.messages)
    word_count = len(normalized.split())
    meta = {
        "intent": intent,
        "topic": topic,
        "feedback_signal": feedback_signal,
        "feedback_reason": feedback_reason,
        "whatsapp_handoff": whatsapp_offer,
    }

    if analysis.theme_slug in {"name", "current-role", "education", "previous-experience", "specific-experience", "tools"} and analysis.reply_text:
        grounded_path = f"chat-grounded-{analysis.theme_slug}"
        return analysis.reply_text, grounded_path, meta

    if detect_conversation_close_intent(normalized):
        close_meta = dict(meta)
        close_meta["whatsapp_handoff"] = True
        return build_conversation_close_reply(language), "chat-close-conversation", close_meta

    if feedback_signal == "negative" and topic == "general":
        return build_feedback_recovery_reply(language, intent, whatsapp_offer, whatsapp_reason), "chat-feedback-recovery", meta

    if is_scope_gap_question(normalized, topic) and not guided_mode:
        scope_path = "chat-scope-gap"
        scope_meta = dict(meta)
        scope_meta["whatsapp_handoff"] = True
        if whatsapp_offer and whatsapp_reason:
            scope_path = f"{scope_path}-whatsapp-{whatsapp_reason}"
        return build_scope_gap_reply(language, True, whatsapp_reason), scope_path, scope_meta

    static_reply = build_static_chat_reply(language, topic, intent, contact_ready, guided_mode, whatsapp_offer, whatsapp_reason)
    if static_reply and should_use_repetition_fallback(static_reply, submission.messages, topic):
        repetition_meta = dict(meta)
        repetition_meta["whatsapp_handoff"] = True
        return build_scope_gap_reply(language, True, whatsapp_reason), "chat-repetition-fallback", repetition_meta
    if static_reply and (topic != "general" or word_count <= 8):
        path = f"chat-static-{topic}" if topic != "general" else "chat-static-short"
        if whatsapp_offer and whatsapp_reason:
            path = f"{path}-whatsapp-{whatsapp_reason}"
        return static_reply, path, meta

    if guided_mode and static_reply:
        guided_path = "chat-static-guided"
        if whatsapp_offer and whatsapp_reason:
            guided_path = f"{guided_path}-whatsapp-{whatsapp_reason}"
        return static_reply, guided_path, meta

    model_reply, engine = openai_chat_reply(
        submission,
        analysis,
        language,
        intent,
        topic,
        contact_ready,
        guided_mode,
        whatsapp_offer,
        whatsapp_reason,
        feedback_signal,
        feedback_reason,
    )
    if model_reply:
        if should_use_repetition_fallback(model_reply, submission.messages, topic):
            repetition_meta = dict(meta)
            repetition_meta["whatsapp_handoff"] = True
            return build_scope_gap_reply(language, True, whatsapp_reason), "chat-repetition-fallback", repetition_meta
        return model_reply, engine, meta

    fallback_path = "chat-fallback"
    fallback_meta = dict(meta)
    fallback_meta["whatsapp_handoff"] = True
    if whatsapp_offer and whatsapp_reason:
        fallback_path = f"{fallback_path}-whatsapp-{whatsapp_reason}"
    return fallback_chat_reply(language, intent, contact_ready, True, whatsapp_reason), fallback_path, fallback_meta


def openai_analysis(submission: InboxSubmission) -> tuple[MessageAnalysis, str]:
    grounded = maybe_build_profile_grounded_analysis(submission)
    if grounded is not None:
        return grounded, "profile-grounded"

    client = get_openai_client()
    if client is None:
        return heuristic_analysis(submission)

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "language": {"type": "string", "enum": ["es", "en"]},
            "category": {"type": "string", "enum": ["question", "suggestion", "project inquiry", "bug report", "general feedback"]},
            "priority": {"type": "string", "enum": ["high", "medium", "low"]},
            "summary": {"type": "string"},
            "key_points": {"type": "array", "items": {"type": "string"}, "minItems": 1, "maxItems": 5},
            "theme_label": {"type": "string"},
            "theme_slug": {"type": "string"},
            "thread_title": {"type": "string"},
            "reply_text": {"type": "string"},
            "lead_score": {"type": "integer", "minimum": 1, "maximum": 5},
            "sentiment": {"type": "string", "enum": ["positive", "neutral", "negative"]},
        },
        "required": ["language", "category", "priority", "summary", "key_points", "theme_label", "theme_slug", "thread_title", "reply_text", "lead_score", "sentiment"],
    }

    prompt = f"""
Analyze this inbound portfolio message for a corporate professional focused on operations, reporting, process support, digital workflows, and practical AI. Carlos currently works at The Retail Performance Company (RPC) since November 2025.

Rules:
- Detect whether the message language is Spanish or English.
- reply_text must be in the same language as the user.
- summary and key_points should stay in the same language as the user.
- theme_label and thread_title should be concise, professional English labels for dashboard consistency.
- theme_slug must be stable, lowercase, hyphenated, and suitable for grouping similar topics.
- Project, collaboration, commercial, or consulting opportunities should usually have a higher lead_score.
- Urgent bugs or time-sensitive project opportunities should often be high priority.
- Do not invent facts.
- Keep summary concise and useful.
- Position Carlos around operations, reporting, process support, dashboards, data visibility, digital workflows, and practical AI.
- Do not describe Carlos as an AI engineer, AI product engineer, or software engineer unless the user explicitly asks about technical implementation details.
- Keep reply_text professional, brief, recruiter-friendly, and realistic.

Metadata:
- sender_name: {submission.name}
- sender_email: {submission.email or "not provided"}
- company: {submission.company or "not provided"}
- source: {submission.source}

Message:
{submission.message}
"""

    try:
        response = client.responses.create(
            model=settings.openai_model,
            input=[
                {"role": "system", "content": [{"type": "input_text", "text": "You classify inbound messages and return strict JSON matching the provided schema."}]},
                {"role": "user", "content": [{"type": "input_text", "text": prompt}]},
            ],
            text={"format": {"type": "json_schema", "name": "bilingual_message_analysis", "strict": True, "schema": schema}},
        )
        parsed = MessageAnalysis.model_validate(json.loads(response.output_text))
        parsed.key_points = normalize_list(parsed.key_points, limit=5)
        parsed.theme_slug = slugify(parsed.theme_slug)
        return parsed, "openai"
    except Exception as exc:
        logger.exception("OpenAI analysis failed, falling back to heuristics: %s", exc)
        return heuristic_analysis(submission)


def text_summary_from_messages(messages: list[dict[str, Any]]) -> str:
    if not messages:
        return "No thread activity yet."
    top_points: list[str] = []
    for message in messages[:5]:
        top_points.extend(message.get("key_points", [])[:2])
    top_points = normalize_list(top_points, limit=4)
    if top_points:
        return "This thread centers on: " + "; ".join(top_points) + "."
    return "This thread contains recurring inbound discussion around a shared topic."


def maybe_generate_ai_thread_summary(messages: list[dict[str, Any]], fallback_summary: str) -> str:
    client = get_openai_client()
    if client is None or not messages:
        return fallback_summary
    snippets = []
    for message in messages[:6]:
        snippets.append(
            f"- language: {message['language']}\n- category: {message['category']}\n- summary: {message['summary']}\n- key points: {', '.join(message['key_points'])}"
        )
    try:
        response = client.responses.create(
            model=settings.openai_model,
            input="Generate a short internal thread summary in English based on these recent messages:\n\n" + "\n\n".join(snippets),
        )
        summary = response.output_text.strip()
        return summary or fallback_summary
    except Exception as exc:
        logger.info("Thread summary fallback used after OpenAI error: %s", exc)
        return fallback_summary


def get_thread_messages(connection: sqlite3.Connection, thread_id: int, limit: int = 6) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            id,
            COALESCE(user_name, sender_name) AS user_name,
            COALESCE(user_email, sender_email) AS user_email,
            company,
            source,
            language,
            category,
            priority,
            lead_score,
            sentiment,
            COALESCE(summary, message_summary) AS summary,
            key_points_json,
            COALESCE(raw_message, message_text) AS raw_message,
            reply_text,
            theme_label,
            theme_slug,
            thread_summary,
            email_status,
            created_at
        FROM messages
        WHERE thread_id = ?
        ORDER BY created_at DESC
        LIMIT ?
        """,
        (thread_id, limit),
    ).fetchall()
    return [
        {
            "id": row["id"],
            "user_name": row["user_name"],
            "user_email": row["user_email"],
            "company": row["company"],
            "source": row["source"],
            "language": row["language"] or "en",
            "category": row["category"] or "general feedback",
            "priority": row["priority"] or "low",
            "lead_score": row["lead_score"] or 1,
            "sentiment": row["sentiment"] or "neutral",
            "summary": row["summary"] or "",
            "key_points": parse_json_list(row["key_points_json"]),
            "raw_message": row["raw_message"] or "",
            "reply_text": row["reply_text"] or "",
            "theme_label": row["theme_label"] or "",
            "theme_slug": row["theme_slug"] or "",
            "thread_summary": row["thread_summary"] or "",
            "email_status": allowed_email_status(row["email_status"] or "skipped"),
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def get_or_create_thread(connection: sqlite3.Connection, analysis: MessageAnalysis) -> int:
    existing = connection.execute(
        "SELECT id FROM threads WHERE theme_slug = ? ORDER BY updated_at DESC, id DESC LIMIT 1",
        (analysis.theme_slug,),
    ).fetchone()
    if existing:
        return int(existing["id"])

    now = utc_now()
    cursor = connection.execute(
        """
        INSERT INTO threads (
            theme_slug, theme_label, thread_title, priority, summary, created_at, updated_at,
            message_count, lead_score, slug, title, category, theme_tags, representative_summary, urgency_score
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, 0)
        """,
        (
            analysis.theme_slug,
            analysis.theme_label,
            analysis.thread_title,
            analysis.priority,
            analysis.summary,
            now,
            now,
            analysis.lead_score,
            analysis.theme_slug,
            analysis.thread_title,
            analysis.category,
            json.dumps([analysis.theme_label]),
            analysis.summary,
        ),
    )
    return int(cursor.lastrowid)


def refresh_thread_rollup(connection: sqlite3.Connection, thread_id: int, analysis: MessageAnalysis) -> dict[str, Any]:
    messages = get_thread_messages(connection, thread_id, limit=6)
    fallback_summary = text_summary_from_messages(messages)
    thread_summary = maybe_generate_ai_thread_summary(messages, fallback_summary)
    highest_priority = "low"
    highest_lead = 1
    for message in messages:
        if PRIORITY_RANK.get(message["priority"], 1) > PRIORITY_RANK.get(highest_priority, 1):
            highest_priority = message["priority"]
        highest_lead = max(highest_lead, int(message["lead_score"]))
    now = utc_now()
    connection.execute(
        """
        UPDATE threads
        SET
            theme_slug = ?,
            theme_label = ?,
            thread_title = ?,
            priority = ?,
            summary = ?,
            updated_at = ?,
            message_count = (SELECT COUNT(*) FROM messages WHERE thread_id = ?),
            lead_score = ?,
            slug = ?,
            title = ?,
            representative_summary = ?,
            theme_tags = ?
        WHERE id = ?
        """,
        (
            analysis.theme_slug,
            analysis.theme_label,
            analysis.thread_title,
            highest_priority,
            thread_summary,
            now,
            thread_id,
            highest_lead,
            analysis.theme_slug,
            analysis.thread_title,
            thread_summary,
            json.dumps([analysis.theme_label]),
            thread_id,
        ),
    )
    connection.execute("UPDATE messages SET thread_summary = ? WHERE thread_id = ?", (thread_summary, thread_id))
    row = connection.execute(
        """
        SELECT
            id,
            theme_slug,
            theme_label,
            COALESCE(thread_title, title) AS thread_title,
            priority,
            COALESCE(summary, representative_summary) AS summary,
            created_at,
            updated_at,
            message_count,
            lead_score
        FROM threads
        WHERE id = ?
        """,
        (thread_id,),
    ).fetchone()
    return dict(row) if row else {}


def allowed_email_status(status: str) -> str:
    normalized = (status or "").strip().lower()
    return normalized if normalized in {"sent", "failed", "skipped"} else "failed"


def smtp_diagnostics() -> dict[str, Any]:
    smtp_ready = all([settings.smtp_host, settings.smtp_username, settings.smtp_password, settings.email_from, settings.owner_email])
    resend_ready = bool(settings.resend_api_key and settings.resend_from_email and settings.owner_email)
    recommendations: list[str] = []

    if not settings.resend_api_key:
        recommendations.append("Falta RESEND_API_KEY. Sin esa clave el backend omitira el envio de emails en Render Free.")
    if not settings.resend_from_email:
        recommendations.append("Falta RESEND_FROM_EMAIL. Debe ser un remitente valido verificado en Resend.")
    if not settings.owner_email:
        recommendations.append("Falta OWNER_EMAIL. Debe ser el correo que recibira las notificaciones.")
    if resend_ready and not recommendations:
        recommendations.append("Configuracion Resend lista para probar el envio real por HTTPS.")

    return {
        "provider": "resend" if resend_ready else "skipped",
        "resend_api_key_configured": bool(settings.resend_api_key),
        "resend_from_email": settings.resend_from_email,
        "resend_ready": resend_ready,
        "smtp_host": settings.smtp_host,
        "smtp_port": settings.smtp_port,
        "smtp_use_tls": settings.smtp_use_tls,
        "smtp_username_configured": bool(settings.smtp_username),
        "smtp_password_configured": bool(settings.smtp_password),
        "smtp_ready": smtp_ready,
        "email_from": settings.email_from,
        "owner_email": settings.owner_email,
        "recomendaciones": recommendations,
    }


def send_email_notification(message_id: int, submission: InboxSubmission, analysis: MessageAnalysis, thread: dict[str, Any], related_messages: list[dict[str, Any]]) -> str:
    diagnostics = smtp_diagnostics()
    logger.info(
        "Email config check | message_id=%s | provider=%s | resend_ready=%s | resend_from_email=%s | owner_email=%s",
        message_id,
        diagnostics["provider"],
        diagnostics["resend_ready"],
        diagnostics["resend_from_email"] or "",
        diagnostics["owner_email"] or "",
    )
    if not diagnostics["resend_ready"]:
        logger.warning("Resend not ready. Message %s stored, email skipped. recomendaciones=%s", message_id, diagnostics["recomendaciones"])
        return "skipped"

    related_lines = []
    for related in related_messages[:3]:
        if related["id"] == message_id:
            continue
        related_lines.append(f"- {related['created_at']} | {related['user_name']} | {related['priority']} | {related['summary']}")

    dashboard_url = f"{settings.app_base_url.rstrip('/')}/dashboard"
    key_points_text = ", ".join(analysis.key_points) if analysis.key_points else "None"
    body = f"""
New AI Portfolio Inbox message

Message metadata
- ID: {message_id}
- Name: {submission.name}
- Email: {submission.email or "not provided"}
- Company: {submission.company or "not provided"}
- Source: {submission.source}
- Language: {analysis.language}
- Category: {analysis.category}
- Priority: {analysis.priority}
- Lead score: {analysis.lead_score}
- Sentiment: {analysis.sentiment}
- Theme label: {analysis.theme_label}
- Thread title: {analysis.thread_title}

AI summary
- Summary: {analysis.summary}
- Key points: {key_points_text}
- Suggested reply: {analysis.reply_text}

Thread context
- Thread summary: {thread.get("summary", "")}
- Thread priority: {thread.get("priority", "")}
- Thread theme slug: {thread.get("theme_slug", "")}
- Message count: {thread.get("message_count", 0)}

Recent related messages
{chr(10).join(related_lines) if related_lines else "- No previous related messages yet."}

Original message
{submission.message}

Dashboard
{dashboard_url}
"""

    subject = f"[Portfolio Inbox] {analysis.priority.upper()} | {analysis.category.title()} | {submission.name}"
    payload = {
        "from": settings.resend_from_email,
        "to": [settings.owner_email],
        "subject": subject,
        "text": body.strip(),
    }
    if submission.email:
        payload["reply_to"] = str(submission.email)

    try:
        logger.info(
            "Resend send start | message_id=%s | to=%s | from=%s | reply_to_present=%s",
            message_id,
            settings.owner_email,
            settings.resend_from_email,
            bool(submission.email),
        )
        response = requests.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {settings.resend_api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=20,
        )
        if response.ok:
            logger.info("Resend send completed | message_id=%s | status_code=%s", message_id, response.status_code)
            return "sent"
        logger.error(
            "Resend send failed | message_id=%s | status_code=%s | body=%s",
            message_id,
            response.status_code,
            response.text[:400],
        )
        return "failed"
    except requests.RequestException as exc:
        logger.exception("Resend delivery failed for message %s: %s", message_id, exc)
        return "failed"
    except Exception:
        logger.exception("Unexpected Resend delivery failure for message %s", message_id)
        return "failed"


def row_value(row: sqlite3.Row, key: str, default: Any = None) -> Any:
    return row[key] if key in row.keys() else default


def serialize_message(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row_value(row, "id"),
        "thread_id": row_value(row, "thread_id"),
        "thread_title": row_value(row, "thread_title", ""),
        "theme_slug": row_value(row, "theme_slug", ""),
        "theme_label": row_value(row, "theme_label", ""),
        "user_name": row_value(row, "user_name", ""),
        "user_email": row_value(row, "user_email", ""),
        "name": row_value(row, "user_name", ""),
        "email": row_value(row, "user_email", ""),
        "company": row_value(row, "company", ""),
        "source": row_value(row, "source", "unknown"),
        "language": row_value(row, "language", "en") or "en",
        "category": row_value(row, "category", "general feedback") or "general feedback",
        "priority": row_value(row, "priority", "low") or "low",
        "lead_score": row_value(row, "lead_score", 1) or 1,
        "sentiment": row_value(row, "sentiment", "neutral") or "neutral",
        "summary": row_value(row, "summary", "") or "",
        "key_points": parse_json_list(row_value(row, "key_points_json", "[]")),
        "raw_message": row_value(row, "raw_message", "") or "",
        "message": row_value(row, "raw_message", "") or "",
        "reply_text": row_value(row, "reply_text", "") or "",
        "thread_summary": row_value(row, "thread_summary", "") or "",
        "email_status": allowed_email_status(row_value(row, "email_status", "skipped") or "skipped"),
        "feedback_signal": row_value(row, "feedback_signal", "none") or "none",
        "feedback_reason": row_value(row, "feedback_reason", "") or "",
        "chat_intent": row_value(row, "chat_intent", "") or "",
        "whatsapp_handoff": bool(row_value(row, "whatsapp_handoff", 0) or 0),
        "created_at": row_value(row, "created_at", utc_now()),
    }


def recent_messages(limit: int = 12) -> list[dict[str, Any]]:
    with closing(get_connection()) as connection:
        rows = connection.execute(
            """
            SELECT
                m.id,
                m.thread_id,
                COALESCE(t.thread_title, t.title) AS thread_title,
                COALESCE(t.theme_slug, m.theme_slug) AS theme_slug,
                COALESCE(t.theme_label, m.theme_label) AS theme_label,
                COALESCE(m.user_name, m.sender_name) AS user_name,
                COALESCE(m.user_email, m.sender_email) AS user_email,
                m.company,
                m.source,
                m.language,
                COALESCE(NULLIF(TRIM(m.category), ''), 'general feedback') AS category,
                m.priority,
                m.lead_score,
                m.sentiment,
                COALESCE(m.summary, m.message_summary) AS summary,
                m.key_points_json,
                COALESCE(m.raw_message, m.message_text) AS raw_message,
                m.reply_text,
                m.thread_summary,
                m.email_status,
                m.created_at
            FROM messages m
            JOIN threads t ON t.id = m.thread_id
            ORDER BY m.created_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    items = [serialize_message(row) for row in rows]
    logger.info(
        "Dashboard recent messages read | database_path=%s | limit=%s | returned=%s",
        settings.database_path,
        limit,
        len(items),
    )
    return items


def safe_recent_messages(limit: int = 12) -> list[dict[str, Any]]:
    try:
        return recent_messages(limit)
    except Exception:
        logger.exception("Recent messages feed failed; returning empty list")
        return []


def thread_rows(limit: int = 20) -> list[dict[str, Any]]:
    with closing(get_connection()) as connection:
        rows = connection.execute(
            """
            SELECT
                id,
                theme_slug,
                theme_label,
                COALESCE(thread_title, title) AS thread_title,
                priority,
                COALESCE(summary, representative_summary) AS summary,
                created_at,
                updated_at,
                message_count,
                lead_score
            FROM threads
            ORDER BY updated_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def aggregate_counts(connection: sqlite3.Connection, column_name: str, fallback: str = "unknown") -> dict[str, int]:
    rows = connection.execute(
        f"""
        SELECT COALESCE(NULLIF(TRIM({column_name}), ''), ?) AS label, COUNT(*) AS total
        FROM messages
        GROUP BY COALESCE(NULLIF(TRIM({column_name}), ''), ?)
        ORDER BY total DESC
        """,
        (fallback, fallback),
    ).fetchall()
    return {row["label"]: row["total"] for row in rows}


def all_messages_for_export(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            m.id,
            m.created_at,
            COALESCE(m.user_name, m.sender_name, 'Website Visitor') AS user_name,
            COALESCE(m.user_email, m.sender_email, '') AS user_email,
            COALESCE(m.company, '') AS company,
            COALESCE(NULLIF(TRIM(m.source), ''), 'unknown') AS source,
            COALESCE(NULLIF(TRIM(m.language), ''), 'unknown') AS language,
            COALESCE(NULLIF(TRIM(m.category), ''), 'other') AS category,
            COALESCE(NULLIF(TRIM(m.priority), ''), 'unknown') AS priority,
            COALESCE(NULLIF(TRIM(m.theme_label), ''), 'Untagged') AS theme_label,
            COALESCE(NULLIF(TRIM(m.theme_slug), ''), 'untagged') AS theme_slug,
            COALESCE(m.summary, m.message_summary, '') AS summary
        FROM messages m
        ORDER BY m.created_at DESC, m.id DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def chart_dataset_from_counts(counts: dict[str, int], limit: int | None = None) -> list[dict[str, Any]]:
    items = [{"label": key or "unknown", "value": int(value)} for key, value in counts.items()]
    items.sort(key=lambda item: (-item["value"], item["label"]))
    return items[:limit] if limit is not None else items


def dashboard_chart_metrics(connection: sqlite3.Connection) -> dict[str, Any]:
    metrics = dashboard_message_metrics(connection)
    by_priority = aggregate_counts(connection, "priority", fallback="unknown")
    by_language = aggregate_counts(connection, "language", fallback="unknown")
    by_category = aggregate_counts(connection, "category", fallback="other")
    by_source = aggregate_counts(connection, "source", fallback="unknown")
    top_themes = [
        {
            "label": item["label"] or "Untagged",
            "slug": item["slug"] or "untagged",
            "value": int(item["total"]),
        }
        for item in metrics["top_themes"]
    ]
    messages_per_day = [
        {
            "label": item["day"] or "unknown",
            "value": int(item["total"]),
        }
        for item in metrics["message_volume"]
    ]
    return {
        "total_messages": metrics["total_messages"],
        "distribution": {
            "priority": chart_dataset_from_counts(by_priority),
            "language": chart_dataset_from_counts(by_language),
            "category": chart_dataset_from_counts(by_category),
            "source": chart_dataset_from_counts(by_source),
        },
        "messages_per_day": messages_per_day,
        "top_themes": top_themes,
        "summary": {
            "by_priority": by_priority,
            "by_language": by_language,
            "by_category": by_category,
            "by_source": by_source,
        },
    }


def dashboard_message_metrics(connection: sqlite3.Connection) -> dict[str, Any]:
    total_messages = connection.execute("SELECT COUNT(*) AS total FROM messages").fetchone()["total"]
    by_source = aggregate_counts(connection, "source", fallback="unknown")
    top_themes = [
        dict(row)
        for row in connection.execute(
            """
            SELECT COALESCE(theme_label, 'General inbound inquiries') AS label,
                   COALESCE(theme_slug, 'general-inquiries') AS slug,
                   COUNT(*) AS total
            FROM messages
            GROUP BY COALESCE(theme_slug, 'general-inquiries'), COALESCE(theme_label, 'General inbound inquiries')
            ORDER BY total DESC
            LIMIT 6
            """
        ).fetchall()
    ]
    message_volume = [
        dict(row)
        for row in connection.execute(
            """
            SELECT substr(created_at, 1, 10) AS day, COUNT(*) AS total
            FROM messages
            WHERE created_at >= ?
            GROUP BY substr(created_at, 1, 10)
            ORDER BY day ASC
            """,
            (iso_days_ago(30),),
        ).fetchall()
    ]
    highest_leads = [
        serialize_message(row)
        for row in connection.execute(
            """
            SELECT
                m.id,
                m.thread_id,
                COALESCE(t.thread_title, t.title) AS thread_title,
                COALESCE(t.theme_slug, m.theme_slug) AS theme_slug,
                COALESCE(t.theme_label, m.theme_label) AS theme_label,
                COALESCE(m.user_name, m.sender_name) AS user_name,
                COALESCE(m.user_email, m.sender_email) AS user_email,
                m.company,
                m.source,
                m.language,
                m.category,
                m.priority,
                m.lead_score,
                m.sentiment,
                COALESCE(m.summary, m.message_summary) AS summary,
                m.key_points_json,
                COALESCE(m.raw_message, m.message_text) AS raw_message,
                m.reply_text,
                m.thread_summary,
                m.email_status,
                m.created_at
            FROM messages m
            JOIN threads t ON t.id = m.thread_id
            ORDER BY m.lead_score DESC, CASE m.priority WHEN 'high' THEN 3 WHEN 'medium' THEN 2 ELSE 1 END DESC, m.created_at DESC
            LIMIT 5
            """
        ).fetchall()
    ]
    recent_high_priority = [
        serialize_message(row)
        for row in connection.execute(
            """
            SELECT
                m.id,
                m.thread_id,
                COALESCE(t.thread_title, t.title) AS thread_title,
                COALESCE(t.theme_slug, m.theme_slug) AS theme_slug,
                COALESCE(t.theme_label, m.theme_label) AS theme_label,
                COALESCE(m.user_name, m.sender_name) AS user_name,
                COALESCE(m.user_email, m.sender_email) AS user_email,
                m.company,
                m.source,
                m.language,
                m.category,
                m.priority,
                m.lead_score,
                m.sentiment,
                COALESCE(m.summary, m.message_summary) AS summary,
                m.key_points_json,
                COALESCE(m.raw_message, m.message_text) AS raw_message,
                m.reply_text,
                m.thread_summary,
                m.email_status,
                m.created_at
            FROM messages m
            JOIN threads t ON t.id = m.thread_id
            WHERE m.priority = 'high'
            ORDER BY m.created_at DESC
            LIMIT 6
            """
        ).fetchall()
    ]
    metrics = {
        "total_messages": total_messages,
        "by_priority": aggregate_counts(connection, "priority"),
        "by_category": aggregate_counts(connection, "category"),
        "by_language": aggregate_counts(connection, "language"),
        "by_sentiment": aggregate_counts(connection, "sentiment"),
        "by_source": by_source,
        "top_themes": top_themes,
        "message_volume": message_volume,
        "highest_leads": highest_leads,
        "recent_high_priority": recent_high_priority,
        "top_opportunities": [item for item in highest_leads if item["lead_score"] >= 4][:4],
        "recurring_interests": top_themes[:4],
    }
    logger.info(
        "Dashboard metrics read | database_path=%s | total_messages=%s | by_source=%s | by_language=%s",
        settings.database_path,
        metrics["total_messages"],
        metrics["by_source"],
        metrics["by_language"],
    )
    return metrics


def fallback_executive_summary(metrics: dict[str, Any]) -> str:
    total = metrics["total_messages"]
    top_theme = metrics["top_themes"][0]["label"] if metrics["top_themes"] else "general inbound inquiries"
    top_opportunities = len(metrics["top_opportunities"])
    return f"The inbox has captured {total} messages so far. The most requested topic is {top_theme}. There are {top_opportunities} strong opportunity signals based on lead score and priority."


def generate_executive_summary(metrics: dict[str, Any]) -> str:
    client = get_openai_client()
    fallback = fallback_executive_summary(metrics)
    if client is None:
        return fallback
    prompt = {
        "total_messages": metrics["total_messages"],
        "by_priority": metrics["by_priority"],
        "by_category": metrics["by_category"],
        "by_language": metrics["by_language"],
        "top_themes": metrics["top_themes"],
        "top_opportunities": [{"summary": item["summary"], "lead_score": item["lead_score"], "priority": item["priority"], "theme_label": item["theme_label"]} for item in metrics["top_opportunities"]],
    }
    try:
        response = client.responses.create(
            model=settings.openai_model,
            input="Write a concise executive summary in English for a portfolio inbox dashboard based on this JSON:\n" + json.dumps(prompt, ensure_ascii=True),
        )
        return response.output_text.strip() or fallback
    except Exception as exc:
        logger.info("Executive summary fallback used after OpenAI error: %s", exc)
        return fallback


def ga4_configured() -> bool:
    return bool(
        settings.ga4_property_id
        and (settings.google_application_credentials or settings.google_application_credentials_json)
        and GA4_CLIENT_AVAILABLE
    )


def get_ga4_client() -> tuple[BetaAnalyticsDataClient, str] | tuple[None, str]:
    if not GA4_CLIENT_AVAILABLE:
        return None, "client_unavailable"
    if settings.google_application_credentials_json:
        try:
            credentials = service_account.Credentials.from_service_account_info(
                json.loads(settings.google_application_credentials_json)
            )
            return BetaAnalyticsDataClient(credentials=credentials), "json_env"
        except Exception:
            logger.exception("Failed to build GA4 client from GOOGLE_APPLICATION_CREDENTIALS_JSON")
            return None, "invalid_json_credentials"
    if settings.google_application_credentials:
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = settings.google_application_credentials
        return BetaAnalyticsDataClient(), "file_path"
    return None, "missing_credentials"


def fetch_ga4_analytics() -> dict[str, Any]:
    if not settings.ga4_property_id:
        logger.warning("GA4 read skipped | reason=missing_property_id")
        return {"status": "not_configured", "reason": "GA4_PROPERTY_ID is not set."}
    if not settings.google_application_credentials and not settings.google_application_credentials_json:
        logger.warning("GA4 read skipped | reason=missing_credentials")
        return {
            "status": "not_configured",
            "reason": "GOOGLE_APPLICATION_CREDENTIALS or GOOGLE_APPLICATION_CREDENTIALS_JSON is not set.",
        }
    if not GA4_CLIENT_AVAILABLE:
        logger.warning("GA4 read skipped | reason=client_library_missing")
        return {"status": "not_configured", "reason": "google-analytics-data client library is not installed."}
    try:
        client, auth_mode = get_ga4_client()
        if client is None:
            return {"status": "error", "reason": f"Unable to initialize GA4 client ({auth_mode})."}
        property_name = f"properties/{settings.ga4_property_id}"
        logger.info(
            "GA4 read start | property=%s | auth_mode=%s",
            property_name,
            auth_mode,
        )
        totals = client.run_report(
            RunReportRequest(
                property=property_name,
                date_ranges=[DateRange(start_date="30daysAgo", end_date="today")],
                metrics=[Metric(name="totalUsers"), Metric(name="sessions"), Metric(name="screenPageViews"), Metric(name="engagedSessions")],
            )
        )
        totals_row = totals.rows[0].metric_values if totals.rows else []
        top_pages_report = client.run_report(
            RunReportRequest(
                property=property_name,
                date_ranges=[DateRange(start_date="30daysAgo", end_date="today")],
                dimensions=[Dimension(name="pagePath")],
                metrics=[Metric(name="screenPageViews")],
                limit=5,
            )
        )
        trend_report = client.run_report(
            RunReportRequest(
                property=property_name,
                date_ranges=[DateRange(start_date="14daysAgo", end_date="today")],
                dimensions=[Dimension(name="date")],
                metrics=[Metric(name="totalUsers"), Metric(name="sessions")],
            )
        )
        channels_report = client.run_report(
            RunReportRequest(
                property=property_name,
                date_ranges=[DateRange(start_date="30daysAgo", end_date="today")],
                dimensions=[Dimension(name="sessionDefaultChannelGroup")],
                metrics=[Metric(name="sessions")],
                limit=6,
            )
        )
        countries_report = client.run_report(
            RunReportRequest(
                property=property_name,
                date_ranges=[DateRange(start_date="7daysAgo", end_date="today")],
                dimensions=[Dimension(name="country")],
                metrics=[Metric(name="activeUsers")],
                limit=8,
            )
        )
        logger.info(
            "GA4 read success | property=%s | users=%s | sessions=%s | page_views=%s",
            property_name,
            int(totals_row[0].value) if len(totals_row) > 0 else 0,
            int(totals_row[1].value) if len(totals_row) > 1 else 0,
            int(totals_row[2].value) if len(totals_row) > 2 else 0,
        )
        return {
            "status": "configured",
            "auth_mode": auth_mode,
            "totals": {
                "users": int(totals_row[0].value) if len(totals_row) > 0 else 0,
                "sessions": int(totals_row[1].value) if len(totals_row) > 1 else 0,
                "page_views": int(totals_row[2].value) if len(totals_row) > 2 else 0,
                "engaged_sessions": int(totals_row[3].value) if len(totals_row) > 3 else 0,
            },
            "top_pages": [{"page": row.dimension_values[0].value or "/", "page_views": int(row.metric_values[0].value)} for row in top_pages_report.rows],
            "time_series": [{"day": row.dimension_values[0].value, "users": int(row.metric_values[0].value), "sessions": int(row.metric_values[1].value)} for row in trend_report.rows],
            "traffic_sources": [{"channel": row.dimension_values[0].value or "Unknown", "sessions": int(row.metric_values[0].value)} for row in channels_report.rows],
            "countries_period_label": "Last 7 days",
            "countries": [
                {
                    "country": row.dimension_values[0].value or "Unknown",
                    "active_users": int(row.metric_values[0].value),
                }
                for row in countries_report.rows
            ],
        }
    except Exception as exc:
        logger.exception("Failed to fetch GA4 analytics: %s", exc)
        return {"status": "error", "reason": str(exc)}


def keyword_overlap_score(label: str, pages: list[dict[str, Any]]) -> int:
    label_tokens = set(slugify(label).split("-"))
    score = 0
    for page in pages:
        page_tokens = set(slugify(page.get("page", "")).split("-"))
        if label_tokens & page_tokens:
            score += int(page.get("page_views", 0))
    return score


def build_combined_insights(metrics: dict[str, Any], analytics: dict[str, Any]) -> dict[str, Any]:
    insights: list[str] = []
    opportunities: list[str] = []
    for item in metrics["top_opportunities"][:3]:
        opportunities.append(f"{item['theme_label']} is showing opportunity potential with lead score {item['lead_score']} and {item['priority']} priority.")
    if analytics.get("status") == "configured":
        for theme in metrics["top_themes"][:3]:
            overlap = keyword_overlap_score(theme["label"], analytics.get("top_pages", []))
            if overlap > 0:
                insights.append(f"{theme['label']} appears directionally aligned with traffic on related pages, based on keyword overlap with top visited content.")
        message_days = {item["day"]: item["total"] for item in metrics["message_volume"]}
        traffic_days = {item["day"]: item["sessions"] for item in analytics.get("time_series", [])}
        shared_days = set(message_days) & set(traffic_days)
        if shared_days:
            peak_day = max(shared_days, key=lambda day: message_days[day] + traffic_days[day])
            insights.append(f"Traffic and inbound activity both peaked around {peak_day}, which may indicate a content-driven contact moment.")
    else:
        insights.append("GA4 is not configured, so combined insights currently rely only on inbox activity.")
    if metrics["top_themes"]:
        insights.append(f"The most requested topic right now is {metrics['top_themes'][0]['label']}.")
    return {
        "summary": fallback_executive_summary(metrics) + " Combined insights are approximate and based on theme frequency plus available traffic patterns.",
        "insights": insights[:5],
        "top_opportunities": opportunities[:4],
        "most_requested_topics": [item["label"] for item in metrics["top_themes"][:5]],
        "analytics_status": analytics.get("status", "not_configured"),
    }


def build_shell_context(request: Request) -> dict[str, Any]:
    return {
        "request": request,
        "has_openai": bool(settings.openai_api_key),
        "has_smtp": bool(settings.smtp_host and settings.smtp_username and settings.smtp_password),
        "ga4_enabled": ga4_configured(),
    }


def safe_dashboard_state() -> dict[str, Any]:
    return {
        "sample_threads": safe_recent_messages(6),
        "dashboard_status": "No data available yet",
    }


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    context = build_shell_context(request)
    try:
        context.update(safe_dashboard_state())
        return templates.TemplateResponse("index.html", context)
    except Exception:
        logger.exception("Index route failed to render full context; serving minimal fallback context")
        context.update({"sample_threads": [], "dashboard_status": "No data available yet"})
        return templates.TemplateResponse("index.html", context)


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request) -> HTMLResponse:
    context = build_shell_context(request)
    try:
        return templates.TemplateResponse("dashboard.html", context)
    except Exception:
        logger.exception("Dashboard route failed to render full context; serving minimal fallback context")
        return templates.TemplateResponse("dashboard.html", context)


@app.get("/api/messages")
async def list_messages(limit: int = Query(default=12, ge=1, le=100)) -> JSONResponse:
    items = recent_messages(limit)
    with closing(get_connection()) as connection:
        total_messages = connection.execute("SELECT COUNT(*) AS total FROM messages").fetchone()["total"]
    by_source: dict[str, int] = {}
    for item in items:
        source = item.get("source") or "unknown"
        by_source[source] = by_source.get(source, 0) + 1
    logger.info(
        "Dashboard raw messages served | database_path=%s | limit=%s | returned=%s | sources=%s",
        settings.database_path,
        limit,
        len(items),
        by_source,
    )
    return JSONResponse(
        {
            "items": items,
            "sources": by_source,
            "total": total_messages,
            "database_path": str(settings.database_path),
        }
    )


@app.get("/api/threads")
async def list_threads(limit: int = Query(default=20, ge=1, le=100)) -> JSONResponse:
    return JSONResponse({"items": thread_rows(limit)})


@app.get("/api/threads/{thread_id}")
async def get_thread(thread_id: int) -> JSONResponse:
    with closing(get_connection()) as connection:
        thread = connection.execute(
            """
            SELECT
                id,
                theme_slug,
                theme_label,
                COALESCE(thread_title, title) AS thread_title,
                priority,
                COALESCE(summary, representative_summary) AS summary,
                created_at,
                updated_at,
                message_count,
                lead_score
            FROM threads
            WHERE id = ?
            """,
            (thread_id,),
        ).fetchone()
        if thread is None:
            raise HTTPException(status_code=404, detail="Thread not found.")
        messages = get_thread_messages(connection, thread_id, limit=20)
    return JSONResponse({"thread": dict(thread), "messages": messages})


@app.options("/api/inbox")
async def inbox_preflight() -> Response:
    return Response(status_code=204)


@app.post("/api/inbox")
async def create_message(request: Request, background_tasks: BackgroundTasks) -> JSONResponse:
    return await handle_inbox_submission(request, background_tasks)

    # ?? FIX: soportar payload tipo chat (messages[])
    is_chat_request = payload.source == CHAT_WIDGET_SOURCE
    if payload.messages:
        last_user_message = next(
            (item.get("content", "").strip() for item in reversed(payload.messages) if item.get("role") == "user" and item.get("content")),
            "",
        )
        if last_user_message:
            payload.message = last_user_message
    logger.info(
        "Inbox request received | source=%s | name=%s | email_provided=%s | company_provided=%s | message_length=%s",
        payload.source,
        payload.name,
        bool(payload.email),
        bool(payload.company),
        len(payload.message),
    )
    logger.info(
        "Inbox payload summary | payload=%s",
        {
            "name": payload.name,
            "email": payload.email,
            "company": payload.company,
            "source": payload.source,
            "message_preview": payload.message[:160],
        },
    )
    if is_chat_request:
        final_language, language_source = resolve_chat_language_details(payload)
        logger.info(
            "Chat request trace | endpoint=/api/inbox | source=%s | chat_locale=%s | chat_language=%s | chat_language_source=%s | history_items=%s",
            payload.source,
            payload.locale or "none",
            final_language,
            language_source,
            len(payload.messages or []),
        )
        analysis, _ = heuristic_analysis(payload)
        reply_text, engine, chat_meta = generate_chat_reply(payload, analysis, final_language)
        logger.info(
            "Chat response trace | chat_response_path=%s | chat_language=%s | feedback_signal=%s | feedback_reason=%s | chat_intent=%s | whatsapp_handoff=%s",
            engine,
            final_language,
            chat_meta["feedback_signal"],
            chat_meta["feedback_reason"] or "none",
            chat_meta["intent"],
            chat_meta["whatsapp_handoff"],
        )
        analysis = analysis.model_copy(update={"language": final_language, "reply_text": reply_text})
    else:
        analysis, engine = openai_analysis(payload)
        chat_meta = {
            "feedback_signal": "none",
            "feedback_reason": None,
            "intent": "",
            "whatsapp_handoff": False,
        }
    logger.info(
        "Inbox analysis completed | source=%s | engine=%s | theme_slug=%s | category=%s | priority=%s",
        payload.source,
        engine,
        analysis.theme_slug,
        analysis.category,
        analysis.priority,
    )
    now = utc_now()
    with closing(get_connection()) as connection:
        try:
            thread_id = get_or_create_thread(connection, analysis)
            logger.info("Inbox SQLite save starting | thread_id=%s | database_path=%s", thread_id, settings.database_path)
            cursor = connection.execute(
                """
                INSERT INTO messages (
                    thread_id, user_name, user_email, company, source, language, category, priority, lead_score,
                    sentiment, summary, key_points_json, raw_message, reply_text, theme_label, theme_slug,
                    thread_summary, email_status, analysis_engine, created_at, sender_name, sender_email,
                    message_text, message_summary, suggested_reply, urgency_score, themes, needs_follow_up,
                    feedback_signal, feedback_reason, chat_intent, whatsapp_handoff
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    thread_id,
                    payload.name,
                    payload.email,
                    payload.company,
                    payload.source,
                    analysis.language,
                    analysis.category,
                    analysis.priority,
                    analysis.lead_score,
                    analysis.sentiment,
                    analysis.summary,
                    json.dumps(analysis.key_points),
                    payload.message,
                    analysis.reply_text,
                    analysis.theme_label,
                    analysis.theme_slug,
                    "",
                    "pending",
                    engine,
                    now,
                    payload.name,
                    payload.email,
                    payload.message,
                    analysis.summary,
                    analysis.reply_text,
                    PRIORITY_RANK.get(analysis.priority, 1) * 25,
                    json.dumps([analysis.theme_label]),
                    1 if analysis.priority != "low" or analysis.lead_score >= 4 else 0,
                    chat_meta["feedback_signal"],
                    chat_meta["feedback_reason"],
                    chat_meta["intent"],
                    1 if chat_meta["whatsapp_handoff"] else 0,
                ),
            )
            message_id = int(cursor.lastrowid)
            logger.info("Inbox SQLite save completed | message_id=%s | thread_id=%s", message_id, thread_id)
            thread = refresh_thread_rollup(connection, thread_id, analysis)
            logger.info(
                "Inbox thread rollup refreshed | thread_id=%s | message_count=%s | priority=%s",
                thread_id,
                thread.get("message_count"),
                thread.get("priority"),
            )
            related_messages = get_thread_messages(connection, thread_id, limit=4)
            email_status = "skipped"
            try:
                logger.info("Inbox email trigger starting | message_id=%s | thread_id=%s", message_id, thread_id)
                email_status = allowed_email_status(
                    send_email_notification(message_id, payload, analysis, thread, related_messages)
                )
                logger.info("Inbox email trigger finished | message_id=%s | email_status=%s", message_id, email_status)
            except Exception as e:
                logger.exception("Email failed but continuing flow")
            connection.execute("UPDATE messages SET email_status = ? WHERE id = ?", (email_status, message_id))
            logger.info("Inbox email status updated | message_id=%s | email_status=%s", message_id, email_status)
            connection.commit()
        except sqlite3.Error as exc:
            logger.exception("Failed to save inbox message: %s", exc)
            raise HTTPException(status_code=500, detail="Unable to save inbox message right now.") from exc
        except Exception as exc:
            logger.exception("Unexpected inbox submission failure: %s", exc)
            raise HTTPException(status_code=500, detail="Unexpected inbox processing error.") from exc

        row = connection.execute(
            """
            SELECT
                m.id,
                m.thread_id,
                COALESCE(t.thread_title, t.title) AS thread_title,
                COALESCE(t.theme_slug, m.theme_slug) AS theme_slug,
                COALESCE(t.theme_label, m.theme_label) AS theme_label,
                COALESCE(m.user_name, m.sender_name) AS user_name,
                COALESCE(m.user_email, m.sender_email) AS user_email,
                m.company,
                m.source,
                m.language,
                m.category,
                m.priority,
                m.lead_score,
                m.sentiment,
                COALESCE(m.summary, m.message_summary) AS summary,
                m.key_points_json,
                COALESCE(m.raw_message, m.message_text) AS raw_message,
                m.reply_text,
                m.thread_summary,
                m.email_status,
                m.created_at
            FROM messages m
            JOIN threads t ON t.id = m.thread_id
            WHERE m.id = ?
            """,
            (message_id,),
        ).fetchone()
    if row is None:
        logger.error("Inbox submission saved but reload failed | message_id=%s", message_id)
        raise HTTPException(status_code=500, detail="Message saved but could not be reloaded.")
    response_payload = {"ok": True, "analysis_engine": engine, "message": serialize_message(row), "thread": thread, "related_messages": related_messages}
    logger.info(
        "Inbox submission completed | message_id=%s | thread_id=%s | response_keys=%s | message_keys=%s",
        response_payload["message"]["id"],
        response_payload["thread"].get("id"),
        list(response_payload.keys()),
        list(response_payload["message"].keys()),
    )
    return JSONResponse(response_payload, status_code=201)


@app.get("/api/dashboard/summary")
async def dashboard_summary() -> JSONResponse:
    with closing(get_connection()) as connection:
        metrics = dashboard_message_metrics(connection)
    metrics["executive_summary"] = generate_executive_summary(metrics)
    logger.info(
        "Dashboard summary served | total_messages=%s | top_theme=%s",
        metrics["total_messages"],
        metrics["top_themes"][0]["label"] if metrics["top_themes"] else "none",
    )
    return JSONResponse(metrics)


@app.get("/api/dashboard/messages")
async def dashboard_messages() -> JSONResponse:
    with closing(get_connection()) as connection:
        metrics = dashboard_message_metrics(connection)
    logger.info(
        "Dashboard messages served | database_path=%s | recent_high_priority=%s | top_opportunities=%s | message_volume_days=%s | sources=%s",
        settings.database_path,
        len(metrics["recent_high_priority"]),
        len(metrics["top_opportunities"]),
        len(metrics["message_volume"]),
        metrics["by_source"],
    )
    return JSONResponse(
        {
            "recent_high_priority": metrics["recent_high_priority"],
            "highest_leads": metrics["highest_leads"],
            "message_volume": metrics["message_volume"],
            "top_opportunities": metrics["top_opportunities"],
            "most_requested_topics": metrics["recurring_interests"],
            "by_source": metrics["by_source"],
        }
    )


@app.get("/api/dashboard/metrics")
async def dashboard_metrics() -> JSONResponse:
    with closing(get_connection()) as connection:
        metrics = dashboard_chart_metrics(connection)
    logger.info(
        "Dashboard chart metrics served | total_messages=%s | source_buckets=%s",
        metrics["total_messages"],
        len(metrics["distribution"]["source"]),
    )
    return JSONResponse(metrics)


@app.get("/api/dashboard/export.xlsx")
async def dashboard_export_excel() -> StreamingResponse:
    with closing(get_connection()) as connection:
        messages = all_messages_for_export(connection)
        metrics = dashboard_chart_metrics(connection)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = "Portfolio Inbox Dashboard Summary"
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A3"] = "Total messages"
    summary_sheet["B3"] = metrics["total_messages"]

    summary_sections = [
        ("By Priority", metrics["summary"]["by_priority"]),
        ("By Language", metrics["summary"]["by_language"]),
        ("By Category", metrics["summary"]["by_category"]),
        ("By Source", metrics["summary"]["by_source"]),
    ]
    current_row = 5
    for title, values in summary_sections:
        summary_sheet.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        current_row += 1
        summary_sheet.cell(row=current_row, column=1, value="Label").font = Font(bold=True)
        summary_sheet.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
        current_row += 1
        for label, total in values.items():
            summary_sheet.cell(row=current_row, column=1, value=label)
            summary_sheet.cell(row=current_row, column=2, value=total)
            current_row += 1
        current_row += 1

    summary_sheet.cell(row=current_row, column=1, value="Top Themes").font = Font(bold=True)
    current_row += 1
    summary_sheet.cell(row=current_row, column=1, value="Theme").font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
    current_row += 1
    for item in metrics["top_themes"]:
        summary_sheet.cell(row=current_row, column=1, value=item["label"])
        summary_sheet.cell(row=current_row, column=2, value=item["value"])
        current_row += 1

    messages_sheet = workbook.create_sheet("Messages")
    headers = [
        "id",
        "created_at",
        "name",
        "email",
        "company",
        "source",
        "language",
        "category",
        "priority",
        "theme_label",
        "theme_slug",
        "summary",
    ]
    messages_sheet.append(headers)
    for cell in messages_sheet[1]:
        cell.font = Font(bold=True)

    for item in messages:
        messages_sheet.append(
            [
                item["id"],
                item["created_at"],
                item["user_name"],
                item["user_email"],
                item["company"],
                item["source"],
                item["language"],
                item["category"],
                item["priority"],
                item["theme_label"],
                item["theme_slug"],
                item["summary"],
            ]
        )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"portfolio-inbox-dashboard-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/dashboard/analytics")
async def dashboard_analytics() -> JSONResponse:
    analytics = fetch_ga4_analytics()
    logger.info(
        "Dashboard analytics served | status=%s | reason=%s",
        analytics.get("status"),
        analytics.get("reason", "none"),
    )
    return JSONResponse(analytics)


@app.get("/api/dashboard/combined-insights")
async def dashboard_combined_insights() -> JSONResponse:
    with closing(get_connection()) as connection:
        metrics = dashboard_message_metrics(connection)
    analytics = fetch_ga4_analytics()
    return JSONResponse(build_combined_insights(metrics, analytics))


@app.get("/api/debug/email")
async def debug_email() -> JSONResponse:
    return JSONResponse(smtp_diagnostics())


@app.get("/health")
async def healthcheck() -> JSONResponse:
    smtp_info = smtp_diagnostics()
    return JSONResponse(
        {
            "status": "ok",
            "timestamp": utc_now(),
            "openai_configured": bool(settings.openai_api_key),
            "smtp_configured": smtp_info["smtp_ready"],
            "resend_configured": smtp_info["resend_ready"],
            "email_delivery_provider": smtp_info["provider"],
            "ga4_configured": ga4_configured(),
            "smtp_host": smtp_info["smtp_host"],
            "smtp_port": smtp_info["smtp_port"],
            "smtp_use_tls": smtp_info["smtp_use_tls"],
            "email_from": smtp_info["email_from"],
            "resend_from_email": smtp_info["resend_from_email"],
            "owner_email": settings.owner_email,
            "smtp_username_configured": smtp_info["smtp_username_configured"],
            "smtp_password_configured": smtp_info["smtp_password_configured"],
            "resend_api_key_configured": smtp_info["resend_api_key_configured"],
            "database_path": str(settings.database_path),
        }
    )



